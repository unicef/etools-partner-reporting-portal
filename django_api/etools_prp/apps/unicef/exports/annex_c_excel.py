import itertools
import logging
import os
import random
import tempfile
import time

from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD, FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter

from etools_prp.apps.indicator.constants import ValueType
from etools_prp.apps.indicator.models import Disaggregation, IndicatorBlueprint
from etools_prp.apps.indicator.utilities import convert_string_number_to_float
from etools_prp.apps.unicef.exports.utilities import PARTNER_PORTAL_DATE_FORMAT_EXCEL
from etools_prp.apps.unicef.models import ProgressReport

logger = logging.getLogger(__name__)


LOCATION_MAX_ADMINISTRATIVE_LEVEL = 5
MAX_DISAGGREGATION_DIMENSIONS = 3


class ProgressReportsXLSXExporter:

    include_disaggregations = False
    export_to_single_sheet = True

    general_info_headers = [
        'Partner Name',
        'Workspace',
        'PD Reference Number',
        'PD Title',
        'PD Output Title',
        'PD Reporting Period',
        'PD Report Status',
        'PD Report Due Date',
        'PD Report Submission Date',
        'Non-Financial contribution to date',
        'Financial contribution to date',
        'Financial contribution currency',
        'Funds received to date',
        'Challenges/bottlenecks in the reporting period',
        'Proposed way forward',
        'Submitted by',
        'FACE Attachment',
        'Other Attachment 1',
        'Other Attachment 2',
        'Narrative',
        'PD Output progress status',
        'PD Output narrative assessment',
        'PD Indicator Title',
        'PD indicator type',
        'PD UNICEF Indicator Target',
        'Calculation method across location',
        'Calculation method across reporting period',
        'Previous location progress',
    ] + list(itertools.chain(*[
        (
            'Location Admin Level {}'.format(level),
            'Admin Level {} PCode'.format(level)
        ) for level in range(1, LOCATION_MAX_ADMINISTRATIVE_LEVEL + 1)
    ])) + [
        'Achievement in reporting period (total across all locations)',
        'Total cumulative progress',
    ]

    column_widths = []

    def __init__(
            self,
            progress_reports,
            include_disaggregations=None,
            export_to_single_sheet=None,
            request=None,
    ):
        self.progress_reports = progress_reports or list()
        filename = ''.join([
            str(time.time()) + str(random.randint(10000000, 100000000))
        ])
        self.file_path = os.path.join(tempfile.gettempdir(), filename + '.xlsx')
        self.display_name = '[{:%a %-d %b %-H-%M-%S %Y}] Progress Report(s) Summary.xlsx'.format(
            timezone.now()
        )
        self.request = request

        self.workbook = Workbook()

        self.current_sheet = self.workbook.active
        if include_disaggregations is not None:
            self.include_disaggregations = include_disaggregations
        if export_to_single_sheet is not None:
            self.export_to_single_sheet = export_to_single_sheet
        self.sheets = [self.current_sheet, ]
        self.disaggregations_start_column = len(self.general_info_headers)

        self.bold_center_style = NamedStyle(name="Bold and Center")
        self.bold_center_style.font = Font(bold=True)
        self.bold_center_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def get_general_info_row(self, progress_report, location_data):
        indicator_report = location_data.indicator_report
        programme_document = progress_report.programme_document

        partner = programme_document.partner

        try:
            indicator_target = convert_string_number_to_float(indicator_report.reportable.calculated_target)
        except ValueError:
            indicator_target = indicator_report.reportable.calculated_target

        previous_location_progress = location_data.previous_location_progress_value

        if indicator_report.is_percentage:
            indicator_report_value_format = FORMAT_PERCENTAGE
            # fixing percent number format: 0.1 is equivalent to 10% in math, openpyxl works the same way
            # e.g. a calculated value of 88.72 becomes 8872% on xls export, that's why it needs divided by 100
            achievement_in_reporting_period = int(indicator_report.total.get(ValueType.CALCULATED, 0)) / 100
            total_cumulative_progress = int(indicator_report.reportable.achieved.get(
                ValueType.CALCULATED, 0
            )) / 100
        else:
            indicator_report_value_format = None
            achievement_in_reporting_period = indicator_report.total.get(ValueType.VALUE, 0)
            total_cumulative_progress = indicator_report.reportable.achieved.get(
                ValueType.VALUE, 0
            )

        face_attachment = None
        other_attachment1 = None
        other_attachment2 = None

        if progress_report.attachments.filter(type="FACE").exists():
            face_attachment = progress_report.attachments.filter(type="FACE").first()

            try:
                # Evaluate the file url to see if it has actual file
                face_attachment.file.url
            except ValueError:
                face_attachment = None

        if progress_report.attachments.filter(type="Other").exists():
            if progress_report.attachments.filter(type="Other").count() > 1:
                other_attachment1 = progress_report.attachments.filter(type="Other").first()
                other_attachment2 = progress_report.attachments.filter(type="Other").last()

                try:
                    # Evaluate the file url to see if it has actual file
                    other_attachment1.file.url
                except ValueError:
                    other_attachment1 = None

                try:
                    # Evaluate the file url to see if it has actual file
                    other_attachment2.file.url
                except ValueError:
                    other_attachment2 = None
            else:
                other_attachment1 = progress_report.attachments.filter(type="Other").first()

                try:
                    # Evaluate the file url to see if it has actual file
                    other_attachment1.file.url
                except ValueError:
                    other_attachment1 = None

        general_info_row = [
            (partner.title, None),
            (', '.join([workspace.title for workspace in location_data.location.workspaces.all()]), None),
            (programme_document.reference_number, None),
            (programme_document.title, None),
            (indicator_report.reportable.content_object.title, None),
            (progress_report.get_reporting_period(), None),
            (progress_report.get_status_display(), None),
            (progress_report.due_date, PARTNER_PORTAL_DATE_FORMAT_EXCEL),
            (progress_report.submission_date, PARTNER_PORTAL_DATE_FORMAT_EXCEL),
            (progress_report.partner_contribution_to_date, None),
            (progress_report.financial_contribution_to_date, None),
            (progress_report.financial_contribution_currency, None),
            (programme_document.funds_received_to_date, FORMAT_CURRENCY_USD),
            (progress_report.challenges_in_the_reporting_period, None),
            (progress_report.proposed_way_forward, None),
            (progress_report.submitted_by.display_name if progress_report.submitted_by else '', None),
            (face_attachment.file.url if face_attachment else '', None),
            (other_attachment1.file.url if other_attachment1 else '', None),
            (other_attachment2.file.url if other_attachment2 else '', None),
            (progress_report.narrative, None),
            (indicator_report.get_overall_status_display(), None),
            (indicator_report.narrative_assessment, None),
            (indicator_report.reportable.blueprint.title, None),
            (indicator_report.display_type, None),
            (indicator_target * 100 if indicator_report.display_type == IndicatorBlueprint.PERCENTAGE else
             indicator_target, None),
            (indicator_report.calculation_formula_across_locations, None),
            (indicator_report.calculation_formula_across_periods, None),
            (previous_location_progress, indicator_report_value_format),
        ]

        location_info = []

        # Iterate over location admin references:
        location = location_data.location
        while True:
            location_info.append([
                location.name, location.p_code
            ])

            if location.parent:
                location = location.parent
            else:
                break

        for i in range(LOCATION_MAX_ADMINISTRATIVE_LEVEL):
            try:
                location_name, location_p_code = location_info[i]
            except IndexError:
                location_name, location_p_code = None, None
            general_info_row.append((location_name, None))
            general_info_row.append((location_p_code, None))

        general_info_row += [
            (achievement_in_reporting_period, indicator_report_value_format),
            (total_cumulative_progress, indicator_report_value_format),
        ]

        return general_info_row

    def fill_workbook(self):
        if self.progress_reports:
            progress_reports = self.progress_reports.select_related(
                'programme_document'
            ).prefetch_related('indicator_reports')

            if self.export_to_single_sheet:
                self.current_sheet.title = 'PRs Export'
                self.write_progress_reports_to_current_sheet(progress_reports)
            else:
                for progress_report in progress_reports:
                    if not self.current_sheet.max_row == 1:
                        self.current_sheet = self.workbook.create_sheet()
                    self.current_sheet.title = 'PR{} Export'.format(progress_report.pk)
                    self.write_progress_reports_to_current_sheet([progress_report])

        self.workbook.save(self.file_path)

    def get_disaggregations(self, progress_reports=None):
        return Disaggregation.objects.filter(
            reportable__indicator_reports__progress_report__in=progress_reports or self.progress_reports
        ).distinct()

    def write_disaggregation_headers_get_column_map(self, progress_reports=None):
        disaggregations = self.get_disaggregations(progress_reports)

        disaggregation_id_to_options = {}
        disaggregation_value_id_to_name = {}
        disaggregation_value_id_to_type = {}

        for disaggregation in disaggregations:
            ids_to_values = dict(disaggregation.disaggregation_values.values_list('id', 'value'))
            disaggregation_value_id_to_name.update(ids_to_values)
            disaggregation_id_to_options[disaggregation.id] = ids_to_values.keys()
            disaggregation_value_id_to_type.update({
                value_id: disaggregation.name for value_id in ids_to_values.keys()
            })

        disaggregation_value_combinations = []

        for dimensions_limit in range(1, MAX_DISAGGREGATION_DIMENSIONS + 1):
            for combinations in itertools.combinations(disaggregation_id_to_options.values(), dimensions_limit):
                disaggregation_value_combinations += list(itertools.product(*combinations))

        disaggregation_value_combinations = map(tuple, map(sorted, disaggregation_value_combinations))
        disaggregation_value_combinations = sorted(
            disaggregation_value_combinations,
            key=lambda l: (len(l), [disaggregation_value_id_to_type[value_id] for value_id in l])
        )

        combination_to_column = {}
        iter_column = -1
        for combination in disaggregation_value_combinations:
            # Same type combinations needs to be excluded
            type_combinations = list()
            exclude_combination = False
            for _id in combination:
                if disaggregation_value_id_to_type[_id] not in type_combinations:
                    type_combinations.append(disaggregation_value_id_to_type[_id])
                else:
                    exclude_combination = True
                    break
            if exclude_combination:
                continue

            iter_column += 1
            column = iter_column + self.disaggregations_start_column + 1

            headers = sorted([
                '{}: {}'.format(
                    disaggregation_value_id_to_type[_id], disaggregation_value_id_to_name[_id]
                ) for _id in combination
            ])

            cell = self.current_sheet.cell(row=1, column=column, value='\n'.join(headers))
            cell.style = self.bold_center_style

            self.current_sheet.merge_cells(
                start_row=1, start_column=column, end_row=MAX_DISAGGREGATION_DIMENSIONS, end_column=column
            )
            self.current_sheet.column_dimensions[get_column_letter(column)].width = max(map(len, headers)) + 2

            # Combinations retrieved from DB are identified by tuple cast to string
            combination_to_column[str(combination)] = column

        # Totals column - add, style and save to mapping
        totals_column = max(combination_to_column.values() or [self.disaggregations_start_column]) + 1
        cell = self.current_sheet.cell(row=1, column=totals_column, value='Total')
        cell.style = self.bold_center_style
        combination_to_column['()'] = totals_column
        self.current_sheet.merge_cells(
            start_row=1, start_column=totals_column, end_row=MAX_DISAGGREGATION_DIMENSIONS, end_column=totals_column
        )

        return combination_to_column

    def write_header_to_current_sheet(self):
        current_row = 1
        for column, header_text in enumerate(self.general_info_headers):
            self.column_widths.append(len(header_text))
            column += 1  # columns are not 0-indexed...
            cell = self.current_sheet.cell(row=current_row, column=column, value=header_text)

            if self.include_disaggregations:
                self.current_sheet.merge_cells(
                    start_row=current_row,
                    start_column=column,
                    end_row=current_row + MAX_DISAGGREGATION_DIMENSIONS - 1,
                    end_column=column
                )

            cell.style = self.bold_center_style
        return current_row + (MAX_DISAGGREGATION_DIMENSIONS if self.include_disaggregations else 1)

    def write_progress_reports_to_current_sheet(self, progress_reports):
        current_row = self.write_header_to_current_sheet()

        disaggregation_column_map = self.write_disaggregation_headers_get_column_map(progress_reports)

        for progress_report in progress_reports:
            current_row = self.write_indicator_reports_to_current_sheet(
                current_row, progress_report.indicator_reports.all(), disaggregation_column_map
            )

        for column, width in enumerate(self.column_widths):
            column += 1
            self.current_sheet.column_dimensions[get_column_letter(column)].width = width

    def write_indicator_reports_to_current_sheet(self, current_row, indicator_reports, disaggregation_column_map):
        for indicator_report in indicator_reports:
            for location_data in indicator_report.indicator_location_data.all():
                general_info_row = self.get_general_info_row(indicator_report.progress_report, location_data)

                for column, (cell_data, cell_format) in enumerate(general_info_row):
                    try:
                        self.column_widths[column] = max(self.column_widths[column], len(cell_data) + 2)
                    except TypeError:
                        self.column_widths[column] = max(self.column_widths[column], len(str(cell_data)) + 2)
                    column += 1  # columns are not 0-indexed...
                    cell = self.current_sheet.cell(row=current_row, column=column, value=cell_data)
                    if cell_format:
                        cell.number_format = cell_format

                for disaggregation, total_value in location_data.disaggregation.items():
                    combination_column = disaggregation_column_map.get(disaggregation)
                    if combination_column:
                        self.current_sheet.cell(
                            row=current_row, column=combination_column, value=total_value.get(ValueType.VALUE)
                        )
                    else:
                        logger.exception('Location data {} contains unknown disaggregation: {}'.format(
                            location_data.id, disaggregation
                        ))

                current_row += 1

        return current_row

    def cleanup(self):
        os.remove(self.file_path)

    def get_as_response(self, request):
        self.fill_workbook()
        response = HttpResponse()
        response.content_type = self.current_sheet.mime_type
        with open(self.file_path, 'rb') as content:
            response.write(content.read())
        self.cleanup()
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(self.display_name)
        return response


class AnnexCXLSXExporter(ProgressReportsXLSXExporter):

    include_disaggregations = True


class SingleProgressReportsXLSXExporter(ProgressReportsXLSXExporter):

    include_disaggregations = True

    def __init__(self, progress_report, **kwargs):
        progress_reports = ProgressReport.objects.filter(id=progress_report.id)
        super().__init__(progress_reports, **kwargs)
