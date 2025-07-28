import csv
import datetime
import os
import random
import tempfile
from unittest import skip
from unittest.mock import patch

from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files import File
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import Q
from django.urls import reverse

from openpyxl import load_workbook
from rest_framework import status
from unicef_notification.models import Notification

from etools_prp.apps.core.common import (
    FINAL_OVERALL_STATUS,
    INDICATOR_REPORT_STATUS,
    OVERALL_STATUS,
    PD_DOCUMENT_TYPE,
    PD_STATUS,
    PR_ATTACHMENT_TYPES,
    PROGRESS_REPORT_STATUS,
    PRP_ROLE_TYPES,
    REPORTING_TYPES,
)
from etools_prp.apps.core.management.commands._generate_disaggregation_fake_data import generate_3_num_disagg_data
from etools_prp.apps.core.models import Location
from etools_prp.apps.core.tests import factories
from etools_prp.apps.core.tests.base import BaseAPITestCase
from etools_prp.apps.core.tests.factories import (
    faker,
    GroupFactory,
    PartnerFactory,
    PartnerUserFactory,
    ProgressReportIndicatorReportFactory,
    RealmFactory,
    WorkspaceFactory,
)
from etools_prp.apps.indicator.disaggregators import QuantityIndicatorDisaggregator
from etools_prp.apps.indicator.models import IndicatorBlueprint, IndicatorLocationData, IndicatorReport, Reportable
from etools_prp.apps.unicef.models import ProgrammeDocument, ProgressReport, ProgressReportAttachment


def convert_xlsx_to_csv(response):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(response.content)
    xlsx_file = load_workbook(tmp.name)
    xlsx_sheet = xlsx_file[xlsx_file.sheetnames[0]]
    csv_filename = tempfile.NamedTemporaryFile()
    with open(csv_filename.name, "w") as csv_file:
        wr = csv.writer(csv_file)
        for row in xlsx_sheet.values:
            wr.writerow(row)
    return csv_filename


def string_in_download(text, response):
    exists = False
    csv_filename = convert_xlsx_to_csv(response)
    with open(csv_filename.name) as csv_file:
        rd = csv.reader(csv_file)
        for row in rd:
            if text in ",".join(row):
                exists = True
                break
    return exists


class TestProgrammeDocumentListAPIView(BaseAPITestCase):

    def setUp(self):
        self.workspace = factories.WorkspaceFactory()
        self.response_plan = factories.ResponsePlanFactory(workspace=self.workspace)
        self.cluster = factories.ClusterFactory(type='cccm', response_plan=self.response_plan)
        self.carto_table = factories.CartoDBTableFactory()
        self.loc1 = factories.LocationFactory()
        self.loc2 = factories.LocationFactory()
        self.loc1.workspaces.add(self.workspace)
        self.loc2.workspaces.add(self.workspace)
        self.unicef_officer = factories.PersonFactory()
        self.unicef_focal_point = factories.PersonFactory()
        self.partner_focal_point = factories.PersonFactory()
        self.objective = factories.ClusterObjectiveFactory(
            cluster=self.cluster,
            locations=[
                self.loc1,
                self.loc2,
            ]
        )
        self.activity = factories.ClusterActivityFactory(
            cluster_objective=self.objective,
            locations=[
                self.loc1, self.loc2
            ]
        )
        self.partner = factories.PartnerFactory(country_code=faker.country_code())
        self.user = factories.NonPartnerUserFactory()
        self.partner_user = factories.PartnerUserFactory(
            workspace=self.workspace,
            partner=self.partner,
            realms__data=[PRP_ROLE_TYPES.ip_authorized_officer]
        )
        factories.ClusterPRPRoleFactory(user=self.user, workspace=self.workspace, cluster=self.cluster, role=PRP_ROLE_TYPES.cluster_imo)
        self.project = factories.PartnerProjectFactory(
            partner=self.partner,
            clusters=[self.cluster],
            locations=[self.loc1, self.loc2],
        )
        self.p_activity = factories.ClusterActivityPartnerActivityFactory(
            partner=self.partner,
            cluster_activity=self.activity,
        )
        self.project_context = factories.PartnerActivityProjectContextFactory(
            project=self.project,
            activity=self.p_activity,
        )
        self.sample_disaggregation_value_map = {
            "height": ["tall", "medium", "short", "extrashort"],
            "age": ["1-2m", "3-4m", "5-6m", '7-10m', '11-13m', '14-16m'],
            "gender": ["male", "female", "other"],
        }

        blueprint = factories.QuantityTypeIndicatorBlueprintFactory(
            unit=IndicatorBlueprint.NUMBER,
            calculation_formula_across_locations=IndicatorBlueprint.SUM,
            calculation_formula_across_periods=IndicatorBlueprint.SUM,
        )
        self.partneractivity_reportable = factories.QuantityReportableToPartnerActivityProjectContextFactory(
            content_object=self.project_context, blueprint=blueprint
        )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc1,
            reportable=self.partneractivity_reportable,
        )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc2,
            reportable=self.partneractivity_reportable,
        )

        self.pd = factories.ProgrammeDocumentFactory(
            workspace=self.workspace,
            partner=self.partner,
            sections=[factories.SectionFactory(), ],
            unicef_officers=[self.unicef_officer, ],
            unicef_focal_point=[self.unicef_focal_point, ],
            partner_focal_point=[self.partner_focal_point, ]
        )

        for idx in range(2):
            qpr_period = factories.QPRReportingPeriodDatesFactory(programme_document=self.pd)
            pr = factories.ProgressReportFactory(
                start_date=qpr_period.start_date,
                end_date=qpr_period.end_date,
                due_date=qpr_period.due_date,
                report_number=idx + 1,
                report_type=qpr_period.report_type,
                is_final=False,
                programme_document=self.pd,
                submitted_by=self.user,
                submitting_user=self.user,
            )

            factories.ProgressReportAttachmentFactory(
                progress_report=pr,
                type=PR_ATTACHMENT_TYPES.face,
            )

        for idx in range(6):
            hr_period = factories.HRReportingPeriodDatesFactory(programme_document=self.pd)
            pr = factories.ProgressReportFactory(
                start_date=hr_period.start_date,
                end_date=hr_period.end_date,
                due_date=hr_period.due_date,
                report_number=idx + 1,
                report_type=hr_period.report_type,
                is_final=False,
                programme_document=self.pd,
                submitted_by=self.user,
                submitting_user=self.user,
            )

            factories.ProgressReportAttachmentFactory(
                progress_report=pr,
                type=PR_ATTACHMENT_TYPES.face,
            )

        self.cp_output = factories.PDResultLinkFactory(
            programme_document=self.pd,
        )
        self.llo = factories.LowerLevelOutputFactory(
            cp_output=self.cp_output,
        )
        self.llo_reportable = factories.QuantityReportableToLowerLevelOutputFactory(
            content_object=self.llo,
            blueprint=factories.QuantityTypeIndicatorBlueprintFactory(
                unit=IndicatorBlueprint.NUMBER,
                calculation_formula_across_locations=IndicatorBlueprint.SUM,
            )
        )

        self.llo_reportable.disaggregations.clear()
        self.partneractivity_reportable.disaggregations.clear()

        # Create the disaggregations and values in the db for all response plans
        # including one for no response plan as well
        for disagg_name, values in self.sample_disaggregation_value_map.items():
            disagg = factories.IPDisaggregationFactory(name=disagg_name)
            cluster_disagg = factories.DisaggregationFactory(name=disagg_name, response_plan=self.response_plan)

            self.llo_reportable.disaggregations.add(disagg)
            self.partneractivity_reportable.disaggregations.add(cluster_disagg)

            for value in values:
                factories.DisaggregationValueFactory(
                    disaggregation=cluster_disagg,
                    value=value
                )
                factories.DisaggregationValueFactory(
                    disaggregation=disagg,
                    value=value
                )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc1,
            reportable=self.llo_reportable,
        )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc2,
            reportable=self.llo_reportable,
        )

        for _ in range(2):
            factories.ClusterIndicatorReportFactory(
                reportable=self.partneractivity_reportable,
                report_status=INDICATOR_REPORT_STATUS.submitted,
            )

        # Creating Level-3 disaggregation location data for all locations
        generate_3_num_disagg_data(self.partneractivity_reportable, indicator_type="quantity")

        for loc_data in IndicatorLocationData.objects.filter(indicator_report__reportable=self.partneractivity_reportable):
            QuantityIndicatorDisaggregator.post_process(loc_data)

        for pr in self.pd.progress_reports.all():
            factories.ProgressReportIndicatorReportFactory(
                progress_report=pr,
                reportable=self.llo_reportable,
                report_status=INDICATOR_REPORT_STATUS.submitted,
                overall_status=OVERALL_STATUS.met,
            )

        # Creating Level-3 disaggregation location data for all locations
        generate_3_num_disagg_data(self.llo_reportable, indicator_type="quantity")

        for loc_data in IndicatorLocationData.objects.filter(indicator_report__reportable=self.llo_reportable):
            QuantityIndicatorDisaggregator.post_process(loc_data)

        super().setUp()

        # Logging in as Partner AO
        self.client.force_authenticate(self.partner_user)

    def tearDown(self):
        for attachment in ProgressReportAttachment.objects.all():
            attachment.file.delete()
            attachment.delete()

    def test_list_api(self):
        url = reverse(
            'programme-document',
            kwargs={'workspace_id': self.workspace.id})
        response = self.client.get(url, format='json')

        self.assertTrue(status.is_success(response.status_code))
        self.assertEqual(len(response.data['results']), 1)

    @skip
    def test_list_filter_api(self):
        url = reverse(
            'programme-document',
            kwargs={'workspace_id': self.workspace.id})
        response = self.client.get(
            url + "?ref_title=&status=&location=",
            format='json'
        )
        self.assertTrue(status.is_success(response.status_code))
        self.assertEqual(len(response.data['results']), 1)

        document = response.data['results'][0]
        response = self.client.get(
            url + "?ref_title=%s&status=&location=" % document['title'][8:],
            format='json'
        )
        self.assertTrue(status.is_success(response.status_code))
        self.assertEqual(len(response.data['results']), 1)
        self.assertEqual(
            response.data['results'][0]['title'],
            document['title'])

        response = self.client.get(
            url + "?ref_title=&status=%s&location=" % document['status'].lower(),
            format='json'
        )
        self.assertTrue(status.is_success(response.status_code))
        self.assertEqual(
            response.data['results'][0]['status'],
            document['status'])
        self.assertEqual(
            response.data['results'][0]['title'],
            document['title'])
        self.assertEqual(
            response.data['results'][0]['reference_number'],
            document['reference_number'])

        # location filtering
        loc = Location.objects.filter(parent__isnull=True).first()
        response = self.client.get(
            url + "?ref_title=&status=&location=%s" % loc.id,
            format='json'
        )

        self.assertTrue(status.is_success(response.status_code))
        for result in response.data['results']:
            self.assertEqual(result['title'], document['title'])

    def test_list_api_export_pdf(self):
        url = reverse(
            'programme-document',
            kwargs={'workspace_id': self.workspace.id})
        response = self.client.get(url, data={"export": "pdf"})

        self.assertTrue(status.is_success(response.status_code))

    def test_sort(self):
        for __ in range(10):
            factories.ProgrammeDocumentFactory(
                workspace=self.workspace,
                partner=self.partner,
            )

        url = reverse(
            'programme-document',
            kwargs={'workspace_id': self.workspace.pk},
        )

        pd_qs = ProgrammeDocument.objects.all()
        self.assertTrue(pd_qs.count() > 1)

        pd = pd_qs.order_by("budget").first()
        response = self.client.get(url, data={"sort": "budget"})
        response_first = response.data["results"][0]
        self.assertEqual(response_first["id"], str(pd.pk))

        # attempt reverse
        pd = pd_qs.order_by("-budget").first()
        response = self.client.get(url, data={"sort": "budget.desc"})
        response_first = response.data["results"][0]
        self.assertEqual(response_first["id"], str(pd.pk))


class TestProgrammeDocumentDetailAPIView(BaseAPITestCase):

    def setUp(self):
        self.workspace = factories.WorkspaceFactory()
        self.response_plan = factories.ResponsePlanFactory(workspace=self.workspace)
        self.cluster = factories.ClusterFactory(type='cccm', response_plan=self.response_plan)
        self.carto_table = factories.CartoDBTableFactory()
        self.loc1 = factories.LocationFactory()
        self.loc2 = factories.LocationFactory()
        self.loc1.workspaces.add(self.workspace)
        self.loc2.workspaces.add(self.workspace)
        self.unicef_officer = factories.PersonFactory()
        self.unicef_focal_point = factories.PersonFactory()
        self.partner_focal_point = factories.PersonFactory()
        self.objective = factories.ClusterObjectiveFactory(
            cluster=self.cluster,
            locations=[
                self.loc1,
                self.loc2,
            ]
        )
        self.activity = factories.ClusterActivityFactory(
            cluster_objective=self.objective,
            locations=[
                self.loc1, self.loc2
            ]
        )
        self.partner = factories.PartnerFactory(country_code=faker.country_code())
        self.user = factories.NonPartnerUserFactory()
        self.partner_user = factories.PartnerUserFactory(
            workspace=self.workspace,
            partner=self.partner,
            realms__data=[PRP_ROLE_TYPES.ip_authorized_officer]
        )
        factories.ClusterPRPRoleFactory(user=self.user, workspace=self.workspace, cluster=self.cluster, role=PRP_ROLE_TYPES.cluster_imo)
        self.project = factories.PartnerProjectFactory(
            partner=self.partner,
            clusters=[self.cluster],
            locations=[self.loc1, self.loc2],
        )
        self.p_activity = factories.ClusterActivityPartnerActivityFactory(
            partner=self.partner,
            cluster_activity=self.activity,
        )
        self.project_context = factories.PartnerActivityProjectContextFactory(
            project=self.project,
            activity=self.p_activity,
        )
        self.sample_disaggregation_value_map = {
            "height": ["tall", "medium", "short", "extrashort"],
            "age": ["1-2m", "3-4m", "5-6m", '7-10m', '11-13m', '14-16m'],
            "gender": ["male", "female", "other"],
        }

        blueprint = factories.QuantityTypeIndicatorBlueprintFactory(
            unit=IndicatorBlueprint.NUMBER,
            calculation_formula_across_locations=IndicatorBlueprint.SUM,
            calculation_formula_across_periods=IndicatorBlueprint.SUM,
        )
        self.partneractivity_reportable = factories.QuantityReportableToPartnerActivityProjectContextFactory(
            content_object=self.project_context, blueprint=blueprint
        )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc1,
            reportable=self.partneractivity_reportable,
        )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc2,
            reportable=self.partneractivity_reportable,
        )

        self.pd = factories.ProgrammeDocumentFactory(
            workspace=self.workspace,
            partner=self.partner,
            sections=[factories.SectionFactory(), ],
            unicef_officers=[self.unicef_officer, ],
            unicef_focal_point=[self.unicef_focal_point, ],
            partner_focal_point=[self.partner_focal_point, ]
        )

        for idx in range(2):
            qpr_period = factories.QPRReportingPeriodDatesFactory(programme_document=self.pd)
            pr = factories.ProgressReportFactory(
                start_date=qpr_period.start_date,
                end_date=qpr_period.end_date,
                due_date=qpr_period.due_date,
                report_number=idx + 1,
                report_type=qpr_period.report_type,
                is_final=False,
                programme_document=self.pd,
                submitted_by=self.user,
                submitting_user=self.user,
            )

            factories.ProgressReportAttachmentFactory(
                progress_report=pr,
                type=PR_ATTACHMENT_TYPES.face,
            )

        for idx in range(6):
            hr_period = factories.HRReportingPeriodDatesFactory(programme_document=self.pd)
            pr = factories.ProgressReportFactory(
                start_date=hr_period.start_date,
                end_date=hr_period.end_date,
                due_date=hr_period.due_date,
                report_number=idx + 1,
                report_type=hr_period.report_type,
                is_final=False,
                programme_document=self.pd,
                submitted_by=self.user,
                submitting_user=self.user,
            )

            factories.ProgressReportAttachmentFactory(
                progress_report=pr,
                type=PR_ATTACHMENT_TYPES.face,
            )

        self.cp_output = factories.PDResultLinkFactory(
            programme_document=self.pd,
        )
        self.llo = factories.LowerLevelOutputFactory(
            cp_output=self.cp_output,
        )
        self.llo_reportable = factories.QuantityReportableToLowerLevelOutputFactory(
            content_object=self.llo,
            blueprint=factories.QuantityTypeIndicatorBlueprintFactory(
                unit=IndicatorBlueprint.NUMBER,
                calculation_formula_across_locations=IndicatorBlueprint.SUM,
            )
        )

        self.llo_reportable.disaggregations.clear()
        self.partneractivity_reportable.disaggregations.clear()

        # Create the disaggregations and values in the db for all response plans
        # including one for no response plan as well
        for disagg_name, values in self.sample_disaggregation_value_map.items():
            disagg = factories.IPDisaggregationFactory(name=disagg_name)
            cluster_disagg = factories.DisaggregationFactory(name=disagg_name, response_plan=self.response_plan)

            self.llo_reportable.disaggregations.add(disagg)
            self.partneractivity_reportable.disaggregations.add(cluster_disagg)

            for value in values:
                factories.DisaggregationValueFactory(
                    disaggregation=cluster_disagg,
                    value=value
                )
                factories.DisaggregationValueFactory(
                    disaggregation=disagg,
                    value=value
                )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc1,
            reportable=self.llo_reportable,
        )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc2,
            reportable=self.llo_reportable,
        )

        for _ in range(2):
            factories.ClusterIndicatorReportFactory(
                reportable=self.partneractivity_reportable,
                report_status=INDICATOR_REPORT_STATUS.submitted,
            )

        # Creating Level-3 disaggregation location data for all locations
        generate_3_num_disagg_data(self.partneractivity_reportable, indicator_type="quantity")

        for loc_data in IndicatorLocationData.objects.filter(indicator_report__reportable=self.partneractivity_reportable):
            QuantityIndicatorDisaggregator.post_process(loc_data)

        for pr in self.pd.progress_reports.all():
            factories.ProgressReportIndicatorReportFactory(
                progress_report=pr,
                reportable=self.llo_reportable,
                report_status=INDICATOR_REPORT_STATUS.submitted,
                overall_status=OVERALL_STATUS.met,
            )

        # Creating Level-3 disaggregation location data for all locations
        generate_3_num_disagg_data(self.llo_reportable, indicator_type="quantity")

        for loc_data in IndicatorLocationData.objects.filter(indicator_report__reportable=self.llo_reportable):
            QuantityIndicatorDisaggregator.post_process(loc_data)

        super().setUp()

        # Logging in as Partner AO
        self.client.force_authenticate(self.partner_user)

    def tearDown(self):
        for attachment in ProgressReportAttachment.objects.all():
            attachment.file.delete()
            attachment.delete()

    def test_detail_api(self):
        url = reverse(
            'programme-document-details',
            kwargs={'pd_id': self.pd.id, 'workspace_id': self.workspace.id})
        response = self.client.get(url, format='json')

        self.assertEqual(len(response.data['locations']), 2)

        actual = [str(loc["id"]) for loc in response.data['locations']]

        self.assertIn(str(self.loc1.id), actual)
        self.assertIn(str(self.loc2.id), actual)

        self.assertTrue(status.is_success(response.status_code))
        self.assertEqual(self.pd.agreement, response.data['agreement'])
        self.assertEqual(
            self.pd.reference_number,
            response.data['reference_number'])


class BaseProgressReportAPITestCase(BaseAPITestCase):

    def setUp(self):
        self.workspace = factories.WorkspaceFactory()
        self.response_plan = factories.ResponsePlanFactory(workspace=self.workspace)
        self.cluster = factories.ClusterFactory(type='cccm', response_plan=self.response_plan)
        self.carto_table = factories.CartoDBTableFactory()
        self.loc1 = factories.LocationFactory()
        self.loc2 = factories.LocationFactory()
        self.loc1.workspaces.add(self.workspace)
        self.loc2.workspaces.add(self.workspace)
        self.unicef_officer = factories.PersonFactory()
        self.unicef_focal_point = factories.PersonFactory()
        self.partner_focal_point = factories.PersonFactory()
        self.government_focal_point = factories.PersonFactory()
        self.objective = factories.ClusterObjectiveFactory(
            cluster=self.cluster,
            locations=[
                self.loc1,
                self.loc2,
            ]
        )
        self.activity = factories.ClusterActivityFactory(
            cluster_objective=self.objective,
            locations=[
                self.loc1, self.loc2
            ]
        )
        self.partner = factories.PartnerFactory()
        self.government = factories.PartnerFactory()

        self.user = factories.NonPartnerUserFactory()
        self.partner_user = factories.PartnerUserFactory(
            workspace=self.workspace,
            partner=self.partner,
            realms__data=[PRP_ROLE_TYPES.ip_authorized_officer]
        )
        RealmFactory(
            user=self.partner_user,
            workspace=self.workspace,
            partner=self.government,
            group=GroupFactory(name=PRP_ROLE_TYPES.ip_authorized_officer)
        )
        factories.ClusterPRPRoleFactory(user=self.user, workspace=self.workspace, cluster=self.cluster, role=PRP_ROLE_TYPES.cluster_imo)
        self.project = factories.PartnerProjectFactory(
            partner=self.partner,
            clusters=[self.cluster],
            locations=[self.loc1, self.loc2],
        )
        self.p_activity = factories.ClusterActivityPartnerActivityFactory(
            partner=self.partner,
            cluster_activity=self.activity,
        )
        self.project_context = factories.PartnerActivityProjectContextFactory(
            project=self.project,
            activity=self.p_activity,
        )
        self.sample_disaggregation_value_map = {
            "height": ["tall", "medium", "short", "extrashort"],
            "age": ["1-2m", "3-4m", "5-6m", '7-10m', '11-13m', '14-16m'],
            "gender": ["male", "female", "other"],
        }

        blueprint = factories.QuantityTypeIndicatorBlueprintFactory(
            unit=IndicatorBlueprint.NUMBER,
            calculation_formula_across_locations=IndicatorBlueprint.SUM,
            calculation_formula_across_periods=IndicatorBlueprint.SUM,
        )
        self.partneractivity_reportable = factories.QuantityReportableToPartnerActivityProjectContextFactory(
            content_object=self.project_context, blueprint=blueprint
        )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc1,
            reportable=self.partneractivity_reportable,
        )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc2,
            reportable=self.partneractivity_reportable,
        )

        self.pd = factories.ProgrammeDocumentFactory(
            workspace=self.workspace,
            partner=self.partner,
            sections=[factories.SectionFactory(), ],
            unicef_officers=[self.unicef_officer, ],
            unicef_focal_point=[self.unicef_focal_point, ],
            partner_focal_point=[self.partner_focal_point, ]
        )

        for idx in range(2):
            qpr_period = factories.QPRReportingPeriodDatesFactory(programme_document=self.pd)
            pr = factories.ProgressReportFactory(
                start_date=qpr_period.start_date,
                end_date=qpr_period.end_date,
                due_date=qpr_period.due_date,
                report_number=idx + 1,
                report_type=qpr_period.report_type,
                is_final=False,
                programme_document=self.pd,
                submitted_by=self.user,
                submitting_user=self.user,
            )

            factories.ProgressReportAttachmentFactory(
                progress_report=pr,
                type=PR_ATTACHMENT_TYPES.face,
            )

        for idx in range(6):
            hr_period = factories.HRReportingPeriodDatesFactory(programme_document=self.pd)
            pr = factories.ProgressReportFactory(
                start_date=hr_period.start_date,
                end_date=hr_period.end_date,
                due_date=hr_period.due_date,
                report_number=idx + 1,
                report_type=hr_period.report_type,
                is_final=False,
                programme_document=self.pd,
                submitted_by=self.user,
                submitting_user=self.user,
            )

            factories.ProgressReportAttachmentFactory(
                progress_report=pr,
                type=PR_ATTACHMENT_TYPES.face,
            )

        self.cp_output = factories.PDResultLinkFactory(
            programme_document=self.pd,
        )
        self.llo = factories.LowerLevelOutputFactory(
            cp_output=self.cp_output,
        )
        self.llo_reportable = factories.QuantityReportableToLowerLevelOutputFactory(
            content_object=self.llo,
            blueprint=factories.QuantityTypeIndicatorBlueprintFactory(
                unit=IndicatorBlueprint.NUMBER,
                calculation_formula_across_locations=IndicatorBlueprint.SUM,
            )
        )

        self.llo_reportable.disaggregations.clear()
        self.partneractivity_reportable.disaggregations.clear()

        # Create the disaggregations and values in the db for all response plans
        # including one for no response plan as well
        for disagg_name, values in self.sample_disaggregation_value_map.items():
            disagg = factories.IPDisaggregationFactory(name=disagg_name)
            cluster_disagg = factories.DisaggregationFactory(name=disagg_name, response_plan=self.response_plan)

            self.llo_reportable.disaggregations.add(disagg)
            self.partneractivity_reportable.disaggregations.add(cluster_disagg)

            for value in values:
                factories.DisaggregationValueFactory(
                    disaggregation=cluster_disagg,
                    value=value
                )
                factories.DisaggregationValueFactory(
                    disaggregation=disagg,
                    value=value
                )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc1,
            reportable=self.llo_reportable,
        )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc2,
            reportable=self.llo_reportable,
        )

        for _ in range(2):
            factories.ClusterIndicatorReportFactory(
                reportable=self.partneractivity_reportable,
                report_status=INDICATOR_REPORT_STATUS.submitted,
            )

        # Creating Level-3 disaggregation location data for all locations
        generate_3_num_disagg_data(self.partneractivity_reportable, indicator_type="quantity")

        for loc_data in IndicatorLocationData.objects.filter(indicator_report__reportable=self.partneractivity_reportable):
            QuantityIndicatorDisaggregator.post_process(loc_data)

        for pr in self.pd.progress_reports.all():
            factories.ProgressReportIndicatorReportFactory(
                progress_report=pr,
                reportable=self.llo_reportable,
                report_status=INDICATOR_REPORT_STATUS.submitted,
                overall_status=OVERALL_STATUS.met,
            )

        # Creating Level-3 disaggregation location data for all locations
        generate_3_num_disagg_data(self.llo_reportable, indicator_type="quantity")

        for loc_data in IndicatorLocationData.objects.filter(indicator_report__reportable=self.llo_reportable):
            QuantityIndicatorDisaggregator.post_process(loc_data)

        # Create GPD  and 3 progress reports
        self.gpd = factories.ProgrammeDocumentFactory(
            document_type=PD_DOCUMENT_TYPE.GDD,
            workspace=self.workspace,
            partner=self.government,
            sections=[factories.SectionFactory(), ],
            unicef_officers=[self.unicef_officer, ],
            unicef_focal_point=[self.unicef_focal_point, ],
            partner_focal_point=[self.government_focal_point, ]
        )

        for idx in range(3):
            qpr_period = factories.QPRReportingPeriodDatesFactory(programme_document=self.gpd)
            pr = factories.ProgressReportFactory(
                start_date=qpr_period.start_date,
                end_date=qpr_period.end_date,
                due_date=qpr_period.due_date,
                report_number=idx + 1,
                report_type=qpr_period.report_type,
                is_final=False,
                programme_document=self.gpd,
                submitted_by=self.user,
                submitting_user=self.user,
            )
            factories.ProgressReportAttachmentFactory(
                progress_report=pr,
                type=PR_ATTACHMENT_TYPES.face,
            )

        super().setUp()

        # Logging in as Partner AO
        self.client.force_authenticate(self.partner_user)

        self.location_id = self.loc1.id
        self.queryset = self.get_queryset()

    def tearDown(self):
        for attachment in ProgressReportAttachment.objects.all():
            attachment.file.delete()
            attachment.delete()

    def get_queryset(self):
        pd_ids = Location.objects.filter(
            Q(id=self.location_id) |
            Q(parent_id=self.location_id) |
            Q(parent__parent_id=self.location_id) |
            Q(parent__parent__parent_id=self.location_id) |
            Q(parent__parent__parent__parent_id=self.location_id)
        ).values_list(
            'reportables__lower_level_outputs__cp_output__programme_document__id',
            flat=True
        )
        return ProgressReport.objects.filter(programme_document_id__in=pd_ids)


class TestProgressReportListAPIView(BaseProgressReportAPITestCase):

    def test_list_api(self):
        url = reverse(
            'progress-reports',
            kwargs={'workspace_id': self.workspace.id})
        response = self.client.get(url, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), self.queryset.count())

    def test_list_api_filter_by_status(self):
        self.reports = self.queryset.filter(
            status=PROGRESS_REPORT_STATUS.due
        )

        url = reverse(
            'progress-reports',
            kwargs={'workspace_id': self.workspace.id})
        url += '?status=' + PROGRESS_REPORT_STATUS.due
        response = self.client.get(url, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), len(self.reports))

    def test_list_api_filter_by_due_status(self):
        self.reports = self.queryset.filter(
            status__in=[
                PROGRESS_REPORT_STATUS.due,
                PROGRESS_REPORT_STATUS.overdue]
        )

        url = reverse(
            'progress-reports',
            kwargs={'workspace_id': self.workspace.id})
        url += '?due=1'
        response = self.client.get(url, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), len(self.reports))

    def test_list_api_filter_by_pd_title(self):
        filter_string = 'reference'
        self.reports = self.queryset.filter(
            Q(programme_document__reference_number__icontains=filter_string) |
            Q(programme_document__title__icontains=filter_string)
        )

        url = reverse(
            'progress-reports',
            kwargs={'workspace_id': self.workspace.id})
        url += '?pd_ref_title=' + filter_string
        response = self.client.get(url, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), len(self.reports))

    def test_list_api_filter_by_due_date(self):
        today = datetime.datetime.today()
        date_format = settings.PRINT_DATA_FORMAT
        pr_ids = ProgressReport.objects.all().values_list('id', flat=True)

        ir_ids = IndicatorReport.objects \
            .filter(progress_report_id__in=pr_ids) \
            .filter(due_date=today) \
            .values_list('progress_report_id') \
            .distinct()
        pr_queryset = self.queryset.filter(indicator_reports__id__in=ir_ids)

        url = reverse(
            'progress-reports',
            kwargs={'workspace_id': self.workspace.id})
        url += '?due_date=' + today.strftime(date_format)
        response = self.client.get(url, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), len(pr_queryset))

    def test_list_api_filter_by_year(self):
        current_year = datetime.datetime.today().year

        pr_queryset = ProgressReport.objects.filter(end_date__year=current_year)
        url = reverse(
            'progress-reports',
            kwargs={'workspace_id': self.workspace.id})
        current_year_url = f'{url}?year={current_year}'
        response = self.client.get(current_year_url, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), len(pr_queryset))

        # test for 0 results on future year
        future_year = (datetime.datetime.today() + datetime.timedelta(days=1365)).year
        future_year_url = f'{url}?year={future_year}'
        response = self.client.get(future_year_url, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 0)

    def test_list_api_filter_by_is_gpd(self):
        self.reports = ProgressReport.objects.filter(programme_document__document_type=PD_DOCUMENT_TYPE.GDD)
        self.partner_user.partner = self.government
        self.partner_user.save()
        url = reverse(
            'progress-reports',
            kwargs={'workspace_id': self.workspace.id})
        url += '?is_gpd=true'
        response = self.client.get(url, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), len(self.reports))

        url = reverse(
            'progress-reports',
            kwargs={'workspace_id': self.workspace.id})
        url += '?is_gpd=false'
        response = self.client.get(url, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 0, 'No PD reports for government context')

    @patch("etools_prp.apps.utils.emails.EmailTemplate.objects.update_or_create")
    @patch.object(Notification, "full_clean", return_value=None)
    @patch.object(Notification, "send_notification", return_value=None)
    def test_list_api_export(self, mock_create, mock_clean, mock_send):
        # ensure at least one report has status submitted
        report = self.queryset.first()
        report.status = PROGRESS_REPORT_STATUS.submitted
        report.save()

        url = reverse(
            'progress-reports',
            kwargs={'workspace_id': self.workspace.id})
        response = self.client.get(url, data={"export": "xlsx"})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        disposition = response.get("Content-Disposition")
        self.assertTrue(disposition.startswith("attachment; filename="))
        self.assertTrue(
            disposition.endswith('Progress Report(s) Summary.xlsx"'),
        )
        self.reports = self.queryset.filter(
            status=PROGRESS_REPORT_STATUS.submitted
        )
        self.assertTrue(len(self.reports))
        self.assertTrue(string_in_download(
            PROGRESS_REPORT_STATUS[PROGRESS_REPORT_STATUS.submitted],
            response,
        ))

    @patch("etools_prp.apps.utils.emails.EmailTemplate.objects.update_or_create")
    @patch.object(Notification, "full_clean", return_value=None)
    @patch.object(Notification, "send_notification", return_value=None)
    def test_list_api_export_filter_empty(self, mock_create, mock_clean, mock_send):
        # ensure we have needed report statuses
        report_overdue = self.queryset.first()
        report_overdue.status = PROGRESS_REPORT_STATUS.overdue
        report_overdue.save()
        report_due = self.queryset.last()
        report_due.status = PROGRESS_REPORT_STATUS.due
        report_due.save()

        url = reverse(
            'progress-reports',
            kwargs={'workspace_id': self.workspace.id})
        response = self.client.get(
            url,
            data={
                "export": "xlsx",
                "status": PROGRESS_REPORT_STATUS.due
            },
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    @patch("etools_prp.apps.utils.emails.EmailTemplate.objects.update_or_create")
    @patch.object(Notification, "full_clean", return_value=None)
    @patch.object(Notification, "send_notification", return_value=None)
    def test_list_api_export_filter(self, mock_create, mock_clean, mock_send):
        # ensure we have needed report statuses
        report_accepted = self.queryset.first()
        report_accepted.status = PROGRESS_REPORT_STATUS.accepted
        report_accepted.save()
        report_submitted = self.queryset.last()
        report_submitted.status = PROGRESS_REPORT_STATUS.submitted
        report_submitted.save()

        url = reverse(
            'progress-reports',
            kwargs={'workspace_id': self.workspace.id})
        response = self.client.get(
            url,
            data={
                "export": "xlsx",
                "status": PROGRESS_REPORT_STATUS.submitted
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        disposition = response.get("Content-Disposition")
        self.assertTrue(disposition.startswith("attachment; filename="))
        self.assertTrue(
            disposition.endswith('Progress Report(s) Summary.xlsx"'),
        )
        reports_submitted = self.queryset.filter(status=PROGRESS_REPORT_STATUS.submitted)
        self.assertTrue(len(reports_submitted))

        reports_accepted = self.queryset.filter(status=PROGRESS_REPORT_STATUS.accepted)
        self.assertTrue(len(reports_accepted))

        self.assertFalse(string_in_download(
            PROGRESS_REPORT_STATUS[PROGRESS_REPORT_STATUS.accepted],
            response,
        ))

    @patch("etools_prp.apps.utils.emails.EmailTemplate.objects.update_or_create")
    @patch.object(Notification, "full_clean", return_value=None)
    @patch.object(Notification, "send_notification", return_value=None)
    def test_list_api_export_filter_multiple(self, mock_create, mock_clean, mock_send):
        # ensure we have needed report statuses
        reports = self.queryset.all()
        self.assertTrue(len(reports) > 3)
        report_overdue = reports[0]
        report_overdue.status = PROGRESS_REPORT_STATUS.overdue
        report_overdue.save()
        report_accepted = reports[1]
        report_accepted.status = PROGRESS_REPORT_STATUS.accepted
        report_accepted.save()
        report_sent_back = reports[2]
        report_sent_back.status = PROGRESS_REPORT_STATUS.sent_back
        report_sent_back.save()

        url = reverse(
            'progress-reports',
            kwargs={'workspace_id': self.workspace.id})
        response = self.client.get(
            url,
            data={
                "export": "xlsx",
                "status": ",".join([
                    PROGRESS_REPORT_STATUS.accepted,
                    PROGRESS_REPORT_STATUS.sent_back,
                ]),
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        disposition = response.get("Content-Disposition")
        self.assertTrue(disposition.startswith("attachment; filename="))
        self.assertTrue(
            disposition.endswith('Progress Report(s) Summary.xlsx"'),
        )
        reports_overdue = self.queryset.filter(
            status=PROGRESS_REPORT_STATUS.overdue
        )
        self.assertTrue(len(reports_overdue))
        reports_accepted = self.queryset.filter(
            status=PROGRESS_REPORT_STATUS.accepted
        )
        self.assertTrue(len(reports_accepted))
        reports_sent_back = self.queryset.filter(
            status=PROGRESS_REPORT_STATUS.sent_back
        )
        self.assertTrue(len(reports_sent_back))
        self.assertFalse(string_in_download(
            PROGRESS_REPORT_STATUS[PROGRESS_REPORT_STATUS.overdue],
            response,
        ))
        self.assertTrue(string_in_download(
            PROGRESS_REPORT_STATUS[PROGRESS_REPORT_STATUS.accepted],
            response,
        ))
        # Only submitted and accepted are allowed to be exported
        self.assertFalse(string_in_download(
            PROGRESS_REPORT_STATUS[PROGRESS_REPORT_STATUS.sent_back],
            response,
        ))

    @patch("etools_prp.apps.utils.emails.EmailTemplate.objects.update_or_create")
    @patch.object(Notification, "full_clean", return_value=None)
    @patch.object(Notification, "send_notification", return_value=None)
    def test_list_api_export_pdf(self, mock_create, mock_clean, mock_send):
        # ensure at least one report has status submitted
        report = self.queryset.first()
        report.status = PROGRESS_REPORT_STATUS.submitted
        report.save()

        url = reverse(
            'progress-reports',
            kwargs={'workspace_id': self.workspace.id})
        response = self.client.get(url, data={"export": "pdf"})

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.reports = self.queryset.filter(
            status=PROGRESS_REPORT_STATUS.submitted
        )
        self.assertTrue(len(self.reports))


class TestProgressReportDetailUpdateAPIView(BaseProgressReportAPITestCase):

    def test_detail_api_not_final(self):
        progress_report = self.pd.progress_reports.filter(is_final=False).first()

        url = reverse(
            'progress-reports-details',
            args=[self.workspace.pk, progress_report.pk],
        )
        response = self.client.get(url, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue('final_review' not in response.data)

    def test_detail_api_final(self):
        progress_report = self.pd.progress_reports.filter(is_final=False).first()
        progress_report.is_final = True
        progress_report.save(update_fields=['is_final'])
        self.assertTrue(hasattr(progress_report, 'final_review'))

        url = reverse(
            'progress-reports-details',
            args=[self.workspace.pk, progress_report.pk],
        )
        response = self.client.get(url, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue('final_review' in response.data)
        self.assertEqual(response.data['final_review']['respond_requests_in_time_choice'], None)
        self.assertEqual(response.data['final_review']['respond_requests_in_time_comment'], None)
        self.assertEqual(response.data['final_review']['respond_requests_in_time_comment'],
                         getattr(progress_report.final_review, 'respond_requests_in_time_comment'))
        for indicator_report in response.data['indicator_reports']:
            self.assertEqual(indicator_report['overall_status_display'], FINAL_OVERALL_STATUS[indicator_report['overall_status']])

    def test_detail_update_not_final(self):
        progress_report = self.pd.progress_reports.filter(is_final=False).first()

        url = reverse(
            'progress-reports-details-update',
            args=[self.workspace.pk, progress_report.pk],
        )
        data = {
            "partner_contribution_to_date": "Partner contribution text",
            "financial_contribution_currency": "USD",
            "challenges_in_the_reporting_period": "Challenges in the reporting period text",
            "proposed_way_forward": "Proposed way forward text"
        }
        response = self.client.put(url, format='json', data=data)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        for field in data.keys():
            self.assertEqual(response.data[field], data[field])

    def test_detail_update_final(self):
        progress_report = self.pd.progress_reports.filter(is_final=False).first()
        progress_report.is_final = True
        progress_report.save(update_fields=['is_final'])

        url = reverse(
            'progress-reports-details-update',
            args=[self.workspace.pk, progress_report.pk],
        )
        data = {
            "financial_contribution_currency": "USD",
            "proposed_way_forward": "Proposed way forward text",
            "final_review": {
                "release_cash_in_time_choice": True,
                "release_cash_in_time_comment": "Unicef did release cash in time",

                "release_supplies_in_time_choice": True,
                "release_supplies_in_time_comment": "Unicef did release supplies in time",

                "feedback_face_form_in_time_choice": False,
                "feedback_face_form_in_time_comment": "Unicef did not feedback in time",

                "respond_requests_in_time_choice": True,
                "respond_requests_in_time_comment": "Unicef did respond in time",

                "implemented_as_planned_choice": False,
                "implemented_as_planned_comment": "Unicef did not implement",

                "action_to_address_choice": True,
                "action_to_address_comment": "Unicef action to address",

                "overall_satisfaction_choice": "very_satisfied",
                "overall_satisfaction_comment": "Very satisfied with Unicef",
            }
        }
        response = self.client.put(url, format='json', data=data)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['proposed_way_forward'], data['proposed_way_forward'])
        progress_report.final_review.refresh_from_db()
        for field in data['final_review'].keys():
            self.assertEqual(response.data['final_review'][field], data['final_review'][field])
            self.assertEqual(response.data['final_review'][field], getattr(progress_report.final_review, field))

    def test_detail_update_gpd_report(self):
        gpd_progress_report = self.gpd.progress_reports.filter(is_final=False).first()
        self.assertFalse(hasattr(gpd_progress_report, 'gpd_report'))

        url = reverse(
            'progress-reports-details-update',
            args=[self.workspace.pk, gpd_progress_report.pk],
        )
        self.partner_user.partner = self.government
        self.partner_user.save()
        data = {
            "delivered_as_planned": "partially",
            "results_achieved": "Results were partially achieved",
            "challenges_in_the_reporting_period": "Challenges in the reporting period text",
            "proposed_way_forward": "Proposed way forward text"
        }
        response = self.client.put(url, format='json', data=data)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        gpd_progress_report.refresh_from_db()
        self.assertTrue(hasattr(gpd_progress_report, 'gpd_report'))
        for field in data.keys():
            self.assertEqual(response.data[field], data[field])
        for field in ["delivered_as_planned", "results_achieved"]:
            self.assertEqual(getattr(gpd_progress_report.gpd_report, field), data[field])

    def test_detail_api_filter_incomplete(self):
        progress_report = self.pd.progress_reports.first()
        ir_qs = IndicatorReport.objects.filter(
            progress_report=progress_report,
        )
        url = reverse(
            'progress-reports-details',
            args=[self.workspace.pk, progress_report.pk],
        )
        response = self.client.get(url, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            len(response.data['indicator_reports']),
            ir_qs.count(),
        )

        ir_qs = ir_qs.filter(submission_date__isnull=True)
        url += '?incomplete=true'
        response = self.client.get(url, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            len(response.data['indicator_reports']),
            ir_qs.count(),
        )

    def test_detail_api_filter_location(self):
        progress_report = self.pd.progress_reports.first()
        ir_qs = IndicatorReport.objects.filter(
            progress_report=progress_report,
        )
        url = reverse(
            'progress-reports-details',
            args=[self.workspace.pk, progress_report.pk],
        )
        response = self.client.get(url, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            len(response.data['indicator_reports']),
            ir_qs.count(),
        )
        new_loc = factories.LocationFactory()
        llo_reportable_2 = factories.QuantityReportableToLowerLevelOutputFactory(
            content_object=self.llo,
            blueprint=factories.QuantityTypeIndicatorBlueprintFactory(
                unit=IndicatorBlueprint.NUMBER,
                calculation_formula_across_locations=IndicatorBlueprint.SUM,
            )
        )
        factories.LocationWithReportableLocationGoalFactory(
            location=new_loc,
            reportable=llo_reportable_2,
        )
        factories.ProgressReportIndicatorReportFactory(
            progress_report=progress_report,
            reportable=llo_reportable_2,
            report_status=INDICATOR_REPORT_STATUS.submitted,
            overall_status=OVERALL_STATUS.met,
        )

        # test an indicator report exists for given location
        ir_qs = ir_qs.filter(reportable__locations__id=new_loc.pk)
        self.assertEqual(ir_qs.count(), 1)
        url_loc_filter = url + f'?location={new_loc.pk}'
        response = self.client.get(url_loc_filter, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            len(response.data['indicator_reports']),
            ir_qs.count(),
        )

        # test no indicator reports exist for an unused location
        unused_loc = factories.LocationFactory()
        ir_qs = ir_qs.filter(reportable__locations__id=unused_loc.pk)
        self.assertEqual(ir_qs.count(), 0)
        url_loc_filter = url + f'?location={unused_loc.pk}'
        response = self.client.get(url_loc_filter, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            len(response.data['indicator_reports']),
            0,
        )

    def test_detail_api_export_pdf(self):
        progress_report = self.pd.progress_reports.first()
        IndicatorReport.objects.filter(
            progress_report=progress_report,
        )
        url = reverse(
            'progress-reports-details',
            args=[self.workspace.pk, progress_report.pk],
        )
        response = self.client.get(url, data={"export": "pdf"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_report_annex_c_export(self):
        progress_report = self.pd.progress_reports.first()
        IndicatorReport.objects.filter(
            progress_report=progress_report,
        )
        url = reverse(
            'progress-reports-pdf',
            args=[self.workspace.pk, progress_report.pk],
        )
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)


class TestGPDProgressReportDetailsUpdateAPIView(BaseProgressReportAPITestCase):

    def test_get_returns_progress_report(self):
        pr = self.pd.progress_reports.first()

        url = reverse(
            "progress-reports-details",
            args=[self.workspace.pk, pr.pk],
        )
        resp = self.client.get(url, format="json")

        self.assertEqual(resp.status_code, status.HTTP_200_OK)

        self.assertIn("partner_contribution_to_date", resp.data)
        self.assertEqual(resp.data["id"], pr.pk)

    def test_put_updates_narrative_fields(self):
        pr = self.pd.progress_reports.filter(is_final=False).first()
        url = reverse(
            "progress-reports-details-update",
            args=[self.workspace.pk, pr.pk],
        )
        payload = {
            "partner_contribution_to_date": "Updated contribution text",
            "financial_contribution_currency": "USD",
            "challenges_in_the_reporting_period": "New challenges text",
            "proposed_way_forward": "New way forward text",
        }
        resp = self.client.put(url, data=payload, format="json")

        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        for field, value in payload.items():
            self.assertEqual(resp.data[field], value)

        # persisted?
        pr.refresh_from_db()
        self.assertEqual(pr.proposed_way_forward, payload["proposed_way_forward"])

    def test_get_includes_indicator_report_summary(self):
        pr = self.pd.progress_reports.first()
        url = reverse(
            "progress-reports-details",
            args=[self.workspace.pk, pr.pk],
        )
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

        irs = resp.data["indicator_reports"]
        self.assertGreaterEqual(len(irs), 1)
        self.assertIn("overall_status", irs[0])


class TestProgressReportPullHFDataAPIView(BaseProgressReportAPITestCase):

    def setUp(self):
        super().setUp()
        self.progress_report = self.pd.progress_reports.filter(
            is_final=False, report_type=REPORTING_TYPES.QPR).first()
        self.hf_indicator_report = self.progress_report.indicator_reports.first()

    def test_get_HF_non_QPR_from_HR_invalid(self):
        self.progress_report.report_type = REPORTING_TYPES.SR
        self.progress_report.save(update_fields=['report_type'])

        url = reverse(
            'progress-reports-pull-hf-data',
            args=[self.workspace.pk, self.progress_report.pk, self.hf_indicator_report.pk],
        )
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue('This Progress Report is not QPR type.' in response.data['non_field_errors'])

    def test_get_HF_from_HR_without_data_invalid(self):
        url = reverse(
            'progress-reports-pull-hf-data',
            args=[self.workspace.pk, self.progress_report.pk, self.hf_indicator_report.pk],
        )
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue('This HR indicator does not have any High frequency reports within this QPR period.' in
                        response.data['non_field_errors'])

    def test_get_HF_from_HR(self):
        hr_progress_report = factories.ProgressReportFactory(
            start_date=self.progress_report.start_date,
            end_date=self.progress_report.end_date,
            due_date=self.progress_report.due_date,
            report_type=REPORTING_TYPES.HR,
            report_number=self.pd.progress_reports.filter(report_type=REPORTING_TYPES.HR).count() + 1,
            is_final=False,
            programme_document=self.pd,
        )
        hf_indicator = ProgressReportIndicatorReportFactory(
            time_period_start=hr_progress_report.start_date,
            time_period_end=hr_progress_report.end_date,
            reportable=self.hf_indicator_report.reportable,
            progress_report=hr_progress_report
        )

        url = reverse(
            'progress-reports-pull-hf-data',
            args=[self.workspace.pk, self.progress_report.pk, self.hf_indicator_report.pk],
        )
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            len(response.data),
            ProgressReport.objects.filter(
                programme_document=self.progress_report.programme_document,
                report_type="HR",
                start_date__gte=self.progress_report.start_date,
                end_date__lte=self.progress_report.end_date,
            ).count()
        )
        hr_ids = [r['id'] for r in response.data]
        self.assertIn(hr_progress_report.pk, hr_ids)
        self.assertEqual(
            hf_indicator.total['v'],
            response.data[hr_ids.index(hr_progress_report.pk)]['report_location_total']['v']
        )

    def test_post_pull_HF_from_HR_unfilled(self):
        hr_progress_report = factories.ProgressReportFactory(
            start_date=self.progress_report.start_date,
            end_date=self.progress_report.end_date,
            due_date=self.progress_report.due_date,
            report_type=REPORTING_TYPES.HR,
            report_number=self.pd.progress_reports.filter(report_type=REPORTING_TYPES.HR).count() + 1,
            is_final=False,
            programme_document=self.pd,
        )
        ProgressReportIndicatorReportFactory(
            time_period_start=hr_progress_report.start_date,
            time_period_end=hr_progress_report.end_date,
            reportable=self.hf_indicator_report.reportable,
            progress_report=hr_progress_report
        )

        url = reverse(
            'progress-reports-pull-hf-data',
            args=[self.workspace.pk, self.progress_report.pk, self.hf_indicator_report.pk],
        )
        response = self.client.post(url)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue(
            "This indicator does not have available data to pull. Enter data for HR report on this indicator first." in
            response.data['non_field_errors'])

    def test_post_pull_HF_from_HR(self):
        hr_progress_report = factories.ProgressReportFactory(
            start_date=self.progress_report.start_date,
            end_date=self.progress_report.end_date,
            due_date=self.progress_report.due_date,
            report_type=REPORTING_TYPES.HR,
            report_number=self.pd.progress_reports.filter(report_type=REPORTING_TYPES.HR).count() + 1,
            is_final=False,
            programme_document=self.pd,
        )
        new_hf_indicator_report = ProgressReportIndicatorReportFactory(
            time_period_start=hr_progress_report.start_date,
            time_period_end=hr_progress_report.end_date,
            reportable=self.hf_indicator_report.reportable,
            progress_report=hr_progress_report
        )
        total = 0
        for idx, loc in enumerate([self.loc1, self.loc2], start=1):
            factories.IndicatorLocationDataFactory(
                indicator_report=new_hf_indicator_report,
                disaggregation={"()": {"c": 0, "d": 0, "v": 100 + idx}},
                disaggregation_reported_on=list(new_hf_indicator_report.disaggregations.values_list(
                    'id', flat=True)),
                location=loc
            )
            total += 100 + idx

        url = reverse(
            'progress-reports-pull-hf-data',
            args=[self.workspace.pk, self.progress_report.pk, self.hf_indicator_report.pk],
        )
        response = self.client.post(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual({'total': {'c': total, 'd': 1, 'v': total}}, response.data)


class TestProgressReportReviewAPIView(BaseProgressReportAPITestCase):

    def test_review_accept_regular_qpr_report(self):
        progress_report = self.pd.progress_reports.filter(
            is_final=False, report_type=REPORTING_TYPES.QPR).first()
        progress_report.programme_document.status = PD_STATUS.active
        progress_report.programme_document.save(update_fields=['status'])
        progress_report.status = PROGRESS_REPORT_STATUS.submitted
        progress_report.save(update_fields=['status'])

        url = reverse(
            'progress-reports-review',
            args=[self.workspace.pk, progress_report.pk],
        )
        default_unicef_user = factories.NonPartnerUserFactory(username=settings.DEFAULT_UNICEF_USER)
        default_unicef_user.jwt_payload = {'email': default_unicef_user.email, 'user_id': default_unicef_user.id}
        data = {
            "status": "Acc",
            "overall_status": OVERALL_STATUS.met,
            "reviewed_by_name": f"{default_unicef_user.first_name} {default_unicef_user.last_name}",
            "review_date": datetime.datetime.now().date()
        }

        self.client.force_authenticate(default_unicef_user)
        response = self.client.post(url, data=data, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        progress_report.refresh_from_db()
        self.assertEqual(progress_report.status, PROGRESS_REPORT_STATUS.accepted)
        self.assertEqual(progress_report.review_overall_status, OVERALL_STATUS.met)
        self.assertEqual(progress_report.review_date, datetime.datetime.now().date())
        self.assertEqual(progress_report.reviewed_by_name, f"{default_unicef_user.first_name} {default_unicef_user.last_name}")
        self.assertEqual(progress_report.reviewed_by_email, default_unicef_user.email)
        self.assertEqual(progress_report.reviewed_by_external_id, default_unicef_user.id)
        # for non-final reports, the met status is mapped to 'Met'
        self.assertEqual(progress_report.indicator_reports.last().get_overall_status_display(), 'Met')

    def test_review_accept_final_qpr_report_invalid(self):
        progress_report = self.pd.progress_reports.filter(
            is_final=False, report_type=REPORTING_TYPES.QPR).first()
        progress_report.programme_document.status = PD_STATUS.active
        progress_report.programme_document.save(update_fields=['status'])
        progress_report.is_final = True
        progress_report.status = PROGRESS_REPORT_STATUS.submitted
        progress_report.save(update_fields=['status', 'is_final'])

        url = reverse(
            'progress-reports-review',
            args=[self.workspace.pk, progress_report.pk],
        )
        default_unicef_user = factories.NonPartnerUserFactory(username=settings.DEFAULT_UNICEF_USER)
        default_unicef_user.jwt_payload = {'email': default_unicef_user.email, 'user_id': default_unicef_user.id}
        data = {
            "status": "Acc",
            "overall_status": OVERALL_STATUS.no_progress,
            "reviewed_by_name": f"{default_unicef_user.first_name} {default_unicef_user.last_name}",
            "review_date": datetime.datetime.now().date()
        }

        self.client.force_authenticate(default_unicef_user)
        response = self.client.post(url, data=data, format='json')

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue('Overall status for a final report is invalid.' in response.data['overall_status'])

    def test_review_accept_final_qpr_report(self):
        progress_report = self.pd.progress_reports.filter(
            is_final=False, report_type=REPORTING_TYPES.QPR).first()
        progress_report.programme_document.status = PD_STATUS.active
        progress_report.programme_document.save(update_fields=['status'])
        progress_report.is_final = True
        progress_report.status = PROGRESS_REPORT_STATUS.submitted
        progress_report.save(update_fields=['status', 'is_final'])

        url = reverse(
            'progress-reports-review',
            args=[self.workspace.pk, progress_report.pk],
        )
        default_unicef_user = factories.NonPartnerUserFactory(username=settings.DEFAULT_UNICEF_USER)
        default_unicef_user.jwt_payload = {'email': default_unicef_user.email, 'user_id': default_unicef_user.id}
        data = {
            "status": "Acc",
            "overall_status": OVERALL_STATUS.met,
            "reviewed_by_name": f"{default_unicef_user.first_name} {default_unicef_user.last_name}",
            "review_date": datetime.datetime.now().date()
        }

        self.client.force_authenticate(default_unicef_user)
        response = self.client.post(url, data=data, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        progress_report.refresh_from_db()
        self.assertEqual(progress_report.status, PROGRESS_REPORT_STATUS.accepted)
        self.assertEqual(progress_report.review_overall_status, OVERALL_STATUS.met)
        self.assertEqual(progress_report.review_date, datetime.datetime.now().date())
        self.assertEqual(progress_report.reviewed_by_name, f"{default_unicef_user.first_name} {default_unicef_user.last_name}")
        self.assertEqual(progress_report.reviewed_by_email, default_unicef_user.email)
        self.assertEqual(progress_report.reviewed_by_external_id, default_unicef_user.id)
        # for final reports, the met status is mapped to 'Achieved as planned'
        self.assertEqual(progress_report.indicator_reports.last().get_overall_status_display(), 'Achieved as planned')

    def test_review_accept_final_qpr_report_send_back(self):
        progress_report = self.pd.progress_reports.filter(
            is_final=False, report_type=REPORTING_TYPES.QPR).first()
        progress_report.programme_document.status = PD_STATUS.active
        progress_report.programme_document.save(update_fields=['status'])
        progress_report.is_final = True
        progress_report.status = PROGRESS_REPORT_STATUS.submitted
        progress_report.save(update_fields=['status', 'is_final'])

        url = reverse(
            'progress-reports-review',
            args=[self.workspace.pk, progress_report.pk],
        )
        default_unicef_user = factories.NonPartnerUserFactory(username=settings.DEFAULT_UNICEF_USER)
        default_unicef_user.jwt_payload = {'email': default_unicef_user.email, 'user_id': default_unicef_user.id}
        data = {
            "status": "Sen",
            "comment": "Comment when sending back",
            "reviewed_by_name": f"{default_unicef_user.first_name} {default_unicef_user.last_name}",
            "review_date": datetime.datetime.now().date()
        }

        self.client.force_authenticate(default_unicef_user)
        response = self.client.post(url, data=data, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        progress_report.refresh_from_db()
        self.assertEqual(progress_report.status, PROGRESS_REPORT_STATUS.sent_back)
        self.assertEqual(progress_report.sent_back_feedback, data['comment'])
        self.assertEqual(progress_report.review_date, datetime.datetime.now().date())
        self.assertEqual(progress_report.reviewed_by_name, f"{default_unicef_user.first_name} {default_unicef_user.last_name}")
        self.assertEqual(progress_report.reviewed_by_email, default_unicef_user.email)
        self.assertEqual(progress_report.reviewed_by_external_id, default_unicef_user.id)

    def test_review_accept_sr_report(self):
        progress_report = factories.ProgressReportFactory(
            report_number=random.randint(1, 50),
            report_type=REPORTING_TYPES.SR,
            status=PROGRESS_REPORT_STATUS.submitted,
            programme_document=self.pd,
        )
        progress_report.programme_document.status = PD_STATUS.active
        progress_report.programme_document.save(update_fields=['status'])

        url = reverse(
            'progress-reports-review',
            args=[self.workspace.pk, progress_report.pk],
        )
        default_unicef_user = factories.NonPartnerUserFactory(username=settings.DEFAULT_UNICEF_USER)
        default_unicef_user.jwt_payload = {'email': default_unicef_user.email, 'user_id': default_unicef_user.id}
        data = {
            "status": "Acc",
            "overall_status": "Met",
            "comment": "",
            "reviewed_by_name": f"{default_unicef_user.first_name} {default_unicef_user.last_name}",
            "review_date": datetime.datetime.now().date()
        }

        self.client.force_authenticate(default_unicef_user)
        response = self.client.post(url, data=data, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        progress_report.refresh_from_db()
        self.assertEqual(progress_report.status, PROGRESS_REPORT_STATUS.accepted)
        self.assertEqual(progress_report.accepted_comment, "")
        self.assertEqual(progress_report.review_date, datetime.datetime.now().date())
        self.assertEqual(progress_report.reviewed_by_name, f"{default_unicef_user.first_name} {default_unicef_user.last_name}")
        self.assertEqual(progress_report.reviewed_by_email, default_unicef_user.email)
        self.assertEqual(progress_report.reviewed_by_external_id, default_unicef_user.id)


class TestProgressReportSubmitAPIView(BaseProgressReportAPITestCase):

    def test_submit_hr_report(self):
        progress_report = self.pd.progress_reports.filter(
            is_final=False, report_type=REPORTING_TYPES.HR).first()
        progress_report.programme_document.status = PD_STATUS.active
        progress_report.programme_document.save(update_fields=['status'])
        progress_report.submission_date = None
        progress_report.save(update_fields=['submission_date'])
        url = reverse(
            'progress-reports-submit',
            args=[self.workspace.pk, progress_report.pk],
        )
        authorized_officer = factories.PartnerUserFactory(
            workspace=self.workspace,
            partner=self.partner,
            email=self.unicef_officer.email,
            realms__data=[PRP_ROLE_TYPES.ip_authorized_officer],
        )
        self.client.force_authenticate(authorized_officer)
        response = self.client.post(url, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        progress_report.refresh_from_db()
        self.assertEqual(progress_report.status, PROGRESS_REPORT_STATUS.accepted)
        self.assertEqual(progress_report.submitted_by, authorized_officer)
        self.assertEqual(progress_report.submitting_user, authorized_officer)
        self.assertEqual(progress_report.submission_date, datetime.datetime.now().date())

    def test_submit_qpr_report(self):
        progress_report = self.pd.progress_reports.filter(
            is_final=False, report_type=REPORTING_TYPES.QPR).first()
        progress_report.programme_document.status = PD_STATUS.active
        progress_report.programme_document.save(update_fields=['status'])
        progress_report.submission_date = None
        progress_report.save(update_fields=['submission_date'])
        url = reverse(
            'progress-reports-submit',
            args=[self.workspace.pk, progress_report.pk],
        )
        authorized_officer = factories.PartnerUserFactory(
            workspace=self.workspace,
            partner=self.partner,
            email=self.unicef_officer.email,
            realms__data=[PRP_ROLE_TYPES.ip_authorized_officer],
        )
        self.client.force_authenticate(authorized_officer)
        response = self.client.post(url, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        progress_report.refresh_from_db()
        self.assertEqual(progress_report.status, PROGRESS_REPORT_STATUS.submitted)
        self.assertEqual(progress_report.submitted_by, authorized_officer)
        self.assertEqual(progress_report.submitting_user, authorized_officer)
        self.assertEqual(progress_report.submission_date, datetime.datetime.now().date())


class TestProgressReportSRSubmitAPIView(BaseProgressReportAPITestCase):

    def test_submit_sr_report(self):
        progress_report = factories.ProgressReportFactory(
            report_number=random.randint(1, 50),
            report_type=REPORTING_TYPES.SR,
            is_final=False,
            programme_document=self.pd,
            submission_date=None,
            submitted_by=None,
            submitting_user=None,
        )
        progress_report.programme_document.status = PD_STATUS.active
        progress_report.programme_document.save(update_fields=['status'])

        url = reverse(
            'progress-reports-submit',
            args=[self.workspace.pk, progress_report.pk],
        )
        authorized_officer = factories.PartnerUserFactory(
            workspace=self.workspace,
            partner=self.partner,
            email=self.unicef_officer.email,
            realms__data=[PRP_ROLE_TYPES.ip_authorized_officer],
        )
        self.client.force_authenticate(authorized_officer)
        response = self.client.post(url, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        progress_report.refresh_from_db()
        self.assertEqual(progress_report.status, PROGRESS_REPORT_STATUS.submitted)
        self.assertEqual(progress_report.submitted_by, authorized_officer)
        self.assertEqual(progress_report.submitting_user, authorized_officer)
        self.assertEqual(progress_report.submission_date, datetime.datetime.now().date())


class TestProgressReportAttachmentListCreateAPIView(BaseAPITestCase):

    def setUp(self):
        self.workspace = factories.WorkspaceFactory()
        self.carto_table = factories.CartoDBTableFactory()
        self.loc1 = factories.LocationFactory()
        self.loc2 = factories.LocationFactory()
        self.loc1.workspaces.add(self.workspace)
        self.loc2.workspaces.add(self.workspace)
        self.unicef_officer = factories.PersonFactory()
        self.unicef_focal_point = factories.PersonFactory()
        self.partner_focal_point = factories.PersonFactory()
        self.partner = factories.PartnerFactory(country_code=faker.country_code())
        self.partner_user = factories.PartnerUserFactory(
            workspace=self.workspace,
            partner=self.partner,
            realms__data=[PRP_ROLE_TYPES.ip_authorized_officer]
        )
        self.sample_disaggregation_value_map = {
            "height": ["tall", "medium", "short", "extrashort"],
            "age": ["1-2m", "3-4m", "5-6m", '7-10m', '11-13m', '14-16m'],
            "gender": ["male", "female", "other"],
        }

        self.pd = factories.ProgrammeDocumentFactory(
            workspace=self.workspace,
            partner=self.partner,
            sections=[factories.SectionFactory(), ],
            unicef_officers=[self.unicef_officer, ],
            unicef_focal_point=[self.unicef_focal_point, ],
            partner_focal_point=[self.partner_focal_point, ]
        )

        for idx in range(2):
            qpr_period = factories.QPRReportingPeriodDatesFactory(programme_document=self.pd)
            pr = factories.ProgressReportFactory(
                start_date=qpr_period.start_date,
                end_date=qpr_period.end_date,
                due_date=qpr_period.due_date,
                report_number=idx + 1,
                report_type=qpr_period.report_type,
                is_final=False,
                programme_document=self.pd,
                submitted_by=self.user,
                submitting_user=self.user,
            )

            factories.ProgressReportAttachmentFactory(
                progress_report=pr,
                type=PR_ATTACHMENT_TYPES.face,
            )

        for idx in range(6):
            hr_period = factories.HRReportingPeriodDatesFactory(programme_document=self.pd)
            pr = factories.ProgressReportFactory(
                start_date=hr_period.start_date,
                end_date=hr_period.end_date,
                due_date=hr_period.due_date,
                report_number=idx + 1,
                report_type=hr_period.report_type,
                is_final=False,
                programme_document=self.pd,
                submitted_by=self.user,
                submitting_user=self.user,
            )

            factories.ProgressReportAttachmentFactory(
                progress_report=pr,
                type=PR_ATTACHMENT_TYPES.face,
            )

        self.cp_output = factories.PDResultLinkFactory(
            programme_document=self.pd,
        )
        self.llo = factories.LowerLevelOutputFactory(
            cp_output=self.cp_output,
        )
        self.llo_reportable = factories.QuantityReportableToLowerLevelOutputFactory(
            content_object=self.llo,
            blueprint=factories.QuantityTypeIndicatorBlueprintFactory(
                unit=IndicatorBlueprint.NUMBER,
                calculation_formula_across_locations=IndicatorBlueprint.SUM,
            )
        )

        self.llo_reportable.disaggregations.clear()

        for disagg_name, values in self.sample_disaggregation_value_map.items():
            disagg = factories.IPDisaggregationFactory(name=disagg_name)

            self.llo_reportable.disaggregations.add(disagg)

            for value in values:
                factories.DisaggregationValueFactory(
                    disaggregation=disagg,
                    value=value
                )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc1,
            reportable=self.llo_reportable,
        )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc2,
            reportable=self.llo_reportable,
        )

        for pr in self.pd.progress_reports.all():
            factories.ProgressReportIndicatorReportFactory(
                progress_report=pr,
                reportable=self.llo_reportable,
                report_status=INDICATOR_REPORT_STATUS.submitted,
                overall_status=OVERALL_STATUS.met,
            )

        # Creating Level-3 disaggregation location data for all locations
        generate_3_num_disagg_data(self.llo_reportable, indicator_type="quantity")

        for loc_data in IndicatorLocationData.objects.filter(indicator_report__reportable=self.llo_reportable):
            QuantityIndicatorDisaggregator.post_process(loc_data)

        super().setUp()

        # Logging in as Partner AO
        self.client.force_authenticate(self.partner_user)

        self.location_id = self.loc1.id
        self.pr = self.pd.progress_reports.first()

        settings.MEDIA_ROOT = tempfile.mkdtemp()

    def tearDown(self):
        for attachment in ProgressReportAttachment.objects.all():
            attachment.file.delete()
            attachment.delete()

    def test_list_api(self):
        url = reverse(
            'progress-reports-attachment-list',
            kwargs={'workspace_id': self.workspace.id, 'progress_report_id': self.pr.id})
        response = self.client.get(url, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), self.pr.attachments.count())

    def test_create_api(self):
        url = reverse(
            'progress-reports-attachment-list',
            kwargs={'workspace_id': self.workspace.id, 'progress_report_id': self.pr.id})

        f = open('test.txt', 'w')
        f.write(factories.faker.text())
        f.close()

        file = File(open('test.txt', 'rb'))
        upload_file = SimpleUploadedFile('test', file.read(), content_type="multipart/form-data")

        data = {'type': 'Other', 'path': upload_file}
        response = self.client.post(url, data, format="multipart")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['size'], data['path'].size)

        os.remove('test.txt')


class TestProgressReportAttachmentAPIView(BaseAPITestCase):

    def setUp(self):
        self.workspace = factories.WorkspaceFactory()
        self.carto_table = factories.CartoDBTableFactory()
        self.loc1 = factories.LocationFactory()
        self.loc2 = factories.LocationFactory()
        self.loc1.workspaces.add(self.workspace)
        self.loc2.workspaces.add(self.workspace)
        self.unicef_officer = factories.PersonFactory()
        self.unicef_focal_point = factories.PersonFactory()
        self.partner_focal_point = factories.PersonFactory()
        self.partner = factories.PartnerFactory(country_code=faker.country_code())
        self.partner_user = factories.PartnerUserFactory(
            workspace=self.workspace,
            partner=self.partner,
            realms__data=[PRP_ROLE_TYPES.ip_authorized_officer]
        )
        self.sample_disaggregation_value_map = {
            "height": ["tall", "medium", "short", "extrashort"],
            "age": ["1-2m", "3-4m", "5-6m", '7-10m', '11-13m', '14-16m'],
            "gender": ["male", "female", "other"],
        }

        self.pd = factories.ProgrammeDocumentFactory(
            workspace=self.workspace,
            partner=self.partner,
            sections=[factories.SectionFactory(), ],
            unicef_officers=[self.unicef_officer, ],
            unicef_focal_point=[self.unicef_focal_point, ],
            partner_focal_point=[self.partner_focal_point, ]
        )

        for idx in range(2):
            qpr_period = factories.QPRReportingPeriodDatesFactory(programme_document=self.pd)
            pr = factories.ProgressReportFactory(
                start_date=qpr_period.start_date,
                end_date=qpr_period.end_date,
                due_date=qpr_period.due_date,
                report_number=idx + 1,
                report_type=qpr_period.report_type,
                is_final=False,
                programme_document=self.pd,
                submitted_by=self.user,
                submitting_user=self.user,
            )

            factories.ProgressReportAttachmentFactory(
                progress_report=pr,
                type=PR_ATTACHMENT_TYPES.face,
            )

        for idx in range(6):
            hr_period = factories.HRReportingPeriodDatesFactory(programme_document=self.pd)
            pr = factories.ProgressReportFactory(
                start_date=hr_period.start_date,
                end_date=hr_period.end_date,
                due_date=hr_period.due_date,
                report_number=idx + 1,
                report_type=hr_period.report_type,
                is_final=False,
                programme_document=self.pd,
                submitted_by=self.user,
                submitting_user=self.user,
            )

            factories.ProgressReportAttachmentFactory(
                progress_report=pr,
                type=PR_ATTACHMENT_TYPES.face,
            )

        self.cp_output = factories.PDResultLinkFactory(
            programme_document=self.pd,
        )
        self.llo = factories.LowerLevelOutputFactory(
            cp_output=self.cp_output,
        )
        self.llo_reportable = factories.QuantityReportableToLowerLevelOutputFactory(
            content_object=self.llo,
            blueprint=factories.QuantityTypeIndicatorBlueprintFactory(
                unit=IndicatorBlueprint.NUMBER,
                calculation_formula_across_locations=IndicatorBlueprint.SUM,
            )
        )

        self.llo_reportable.disaggregations.clear()

        for disagg_name, values in self.sample_disaggregation_value_map.items():
            disagg = factories.IPDisaggregationFactory(name=disagg_name)

            self.llo_reportable.disaggregations.add(disagg)

            for value in values:
                factories.DisaggregationValueFactory(
                    disaggregation=disagg,
                    value=value
                )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc1,
            reportable=self.llo_reportable,
        )

        factories.LocationWithReportableLocationGoalFactory(
            location=self.loc2,
            reportable=self.llo_reportable,
        )

        for pr in self.pd.progress_reports.all():
            factories.ProgressReportIndicatorReportFactory(
                progress_report=pr,
                reportable=self.llo_reportable,
                report_status=INDICATOR_REPORT_STATUS.submitted,
                overall_status=OVERALL_STATUS.met,
            )

        # Creating Level-3 disaggregation location data for all locations
        generate_3_num_disagg_data(self.llo_reportable, indicator_type="quantity")

        for loc_data in IndicatorLocationData.objects.filter(indicator_report__reportable=self.llo_reportable):
            QuantityIndicatorDisaggregator.post_process(loc_data)

        super().setUp()

        # Logging in as Partner AO
        self.client.force_authenticate(self.partner_user)

        self.location_id = self.loc1.id
        self.pr = self.pd.progress_reports.first()
        self.attachment = self.pr.attachments.first()

        settings.MEDIA_ROOT = tempfile.mkdtemp()

    def tearDown(self):
        for attachment in ProgressReportAttachment.objects.all():
            attachment.file.delete()
            attachment.delete()

    def test_detail_api(self):
        url = reverse(
            'progress-reports-attachment',
            kwargs={'workspace_id': self.workspace.id, 'progress_report_id': self.pr.id, 'pk': self.attachment.id})
        response = self.client.get(url, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['id'], self.attachment.id)

    def test_update_api(self):
        f = open('test.txt', 'w')
        f.write(factories.faker.text() + factories.faker.text() + factories.faker.text())
        f.close()

        file = File(open('test.txt', 'rb'))
        upload_file = SimpleUploadedFile('test', file.read(), content_type="multipart/form-data")

        data = {'type': 'Other', 'path': upload_file}
        url = reverse(
            'progress-reports-attachment',
            kwargs={'workspace_id': self.workspace.id, 'progress_report_id': self.pr.id, 'pk': self.attachment.id})
        response = self.client.put(url, data, format="multipart")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['size'], data['path'].size)

        os.remove('test.txt')

    def test_delete_api(self):
        url = reverse(
            'progress-reports-attachment',
            kwargs={'workspace_id': self.workspace.id, 'progress_report_id': self.pr.id, 'pk': self.attachment.id})
        response = self.client.delete(url, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)


class TestProgrammeDocumentIndicatorsAPIView(BaseAPITestCase):
    def setUp(self):
        self.workspace = factories.WorkspaceFactory()
        self.partner = factories.PartnerFactory(country_code=faker.country_code())
        self.user = factories.PartnerUserFactory(
            workspace=self.workspace,
            partner=self.partner,
            realms__data=[PRP_ROLE_TYPES.ip_authorized_officer]
        )
        self.unicef_officer = factories.PersonFactory()
        self.unicef_focal_point = factories.PersonFactory()
        self.partner_focal_point = factories.PersonFactory()
        self.pd = factories.ProgrammeDocumentFactory(
            workspace=self.workspace,
            partner=self.partner,
            sections=[factories.SectionFactory(), ],
            unicef_officers=[self.unicef_officer, ],
            unicef_focal_point=[self.unicef_focal_point, ],
            partner_focal_point=[self.partner_focal_point, ]
        )
        self.cp_output = factories.PDResultLinkFactory(
            programme_document=self.pd,
        )
        self.llo = factories.LowerLevelOutputFactory(
            cp_output=self.cp_output,
        )
        self.report_number = 1

        super().setUp()

    def _setup_reportable(self, report_status=None):
        reportable = factories.QuantityReportableToLowerLevelOutputFactory(
            content_object=self.llo,
            blueprint=factories.QuantityTypeIndicatorBlueprintFactory(
                unit=IndicatorBlueprint.NUMBER,
                calculation_formula_across_locations=IndicatorBlueprint.SUM,
            )
        )
        self.rep_loc_goal_1 = factories.LocationWithReportableLocationGoalFactory(
            location=factories.LocationFactory(),
            reportable=reportable,
        )

        self.rep_loc_goal_2 = factories.LocationWithReportableLocationGoalFactory(
            location=factories.LocationFactory(),
            reportable=reportable,
        )

        hr_period = factories.HRReportingPeriodDatesFactory(programme_document=self.pd)
        report_status = report_status if report_status else random.choice(
            [x[0] for x in PROGRESS_REPORT_STATUS]
        )
        pr = factories.ProgressReportFactory(
            start_date=hr_period.start_date,
            end_date=hr_period.end_date,
            due_date=hr_period.due_date,
            report_number=self.report_number,
            report_type=hr_period.report_type,
            is_final=False,
            status=report_status,
            programme_document=self.pd,
            submitted_by=self.user,
            submitting_user=self.user,
        )
        self.report_number += 1
        factories.ProgressReportIndicatorReportFactory(
            progress_report=pr,
            reportable=reportable,
            report_status=INDICATOR_REPORT_STATUS.submitted,
            overall_status=OVERALL_STATUS.met,
        )
        return reportable, pr

    def test_get(self):
        reportable, __ = self._setup_reportable()
        reportable_qs = Reportable.objects.filter(
            lower_level_outputs__isnull=False,
        )
        url = reverse(
            "programme-document-indicators",
            args=[self.workspace.pk],
        )
        self.assertTrue(reportable_qs.count())
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["results"]), reportable_qs.count())
        data = response.data["results"][0]
        self.assertEqual(data["id"], reportable.pk)

    def test_list_without_inactive_reportable_location_goal(self):
        url = reverse(
            "programme-document-indicators",
            args=[self.workspace.pk],
        )
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        actual_locations = [i['location'] for i in response.data["results"][0]['locations']]
        actual_locations.sort()
        expected_locations = [self.rep_loc_goal_1.location.id, self.rep_loc_goal_2.location.id]
        expected_locations.sort()
        self.assertEqual(actual_locations, expected_locations)

        self.rep_loc_goal_1.is_active = False
        self.rep_loc_goal_1.save(update_fields=['is_active'])
        url = reverse(
            "programme-document-indicators",
            args=[self.workspace.pk],
        )
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["results"][0]['locations']), 1)
        self.assertEqual(response.data["results"][0]['locations'][0]['location'], self.rep_loc_goal_2.location.id)

    def test_filter_report_status(self):
        report_status = "Sub"
        reportable, __ = self._setup_reportable(report_status=report_status)
        self.assertEqual(reportable.indicator_reports.count(), 1)
        reportable_qs = Reportable.objects.filter(
            indicator_reports__progress_report__status=report_status
        )
        self.assertTrue(reportable_qs.exists())
        url = reverse(
            "programme-document-indicators",
            args=[self.workspace.pk],
        )

        # expect results
        response = self.client.get(f"{url}?report_status={report_status}")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["results"]), 1)

        # expect no results
        for different_report_status, __ in PROGRESS_REPORT_STATUS:
            if different_report_status != report_status:
                break
        response = self.client.get(
            f"{url}?report_status={different_report_status}"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["results"]), 0)

    def test_export(self):
        report_status = "Sub"
        url = reverse(
            "programme-document-indicators",
            args=[self.workspace.pk],
        )
        reportable, __ = self._setup_reportable(report_status=report_status)

        # expect results
        response = self.client.get(
            f"{url}?report_status={report_status}&export=pdf"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        # add reportable that should not be part of export
        self._setup_reportable(report_status="Due")
        response = self.client.get(
            f"{url}?report_status={report_status}&export=pdf"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)


class TestEToolsRolesSynchronization(BaseAPITestCase):
    @classmethod
    def setUpTestData(cls):
        GroupFactory(name='IP_VIEWER')
        GroupFactory(name='IP_EDITOR')

    def test_sync(self):
        user = PartnerUserFactory(realms__data=['IP_VIEWER'])
        self.assertIsNotNone(user.workspace.external_id)
        self.assertIsNotNone(user.partner.vendor_number)
        group_to_activate = GroupFactory(name='IP_AUTHORIZED_OFFICER')
        RealmFactory(
            user=user,
            workspace=user.workspace,
            partner=user.partner,
            group=group_to_activate,
            is_active=False,
        )
        new_group = GroupFactory(name='IP_EDITOR')
        input_data = {
            'email': user.email,
            'first_name': 'John',
            'middle_name': '_',
            'last_name': 'Doe',
            'realms': [
                {
                    'country': user.workspace.external_id,
                    'organization': user.partner.vendor_number,
                    'group': "IP Editor",
                },
                {
                    'country': user.workspace.external_id,
                    'organization': user.partner.vendor_number,
                    'group': "IP Authorized Officer",
                },
            ]
        }
        self.client.force_authenticate(factories.NonPartnerUserFactory())
        response = self.client.post(reverse('user-realms-import'), data=input_data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        user.refresh_from_db()
        self.assertEqual(user.first_name, 'John')
        self.assertEqual(user.middle_name, '_')
        self.assertEqual(user.last_name, 'Doe')
        self.assertCountEqual(
            list(user.realms.all().values_list('workspace', 'partner', 'group__name', 'is_active')),
            [
                (user.workspace.id, user.partner.id, Group.objects.get(name='IP_VIEWER').name, False),
                (user.workspace.id, user.partner.id, new_group.name, True),
                (user.workspace.id, user.partner.id, group_to_activate.name, True),
            ]
        )

    def test_create_user(self):
        user = PartnerUserFactory.build(realms__data=[])
        workspace = WorkspaceFactory()
        partner = PartnerFactory()
        self.assertFalse(get_user_model().objects.filter(email=user.email).exists())

        input_data = {
            'email': user.email,
            'first_name': user.first_name,
            'middle_name': user.middle_name,
            'last_name': user.last_name,
            'realms': [
                {
                    'country': workspace.external_id,
                    'organization': partner.vendor_number,
                    'group': "IP Editor",
                },
                {
                    'country': workspace.external_id,
                    'organization': partner.vendor_number,
                    'group': "IP Viewer",
                },
            ]
        }
        self.client.force_authenticate(factories.NonPartnerUserFactory())
        response = self.client.post(reverse('user-realms-import'), data=input_data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        self.assertTrue(get_user_model().objects.filter(email=user.email).exists())
        user = get_user_model().objects.get(email=user.email)
        self.assertCountEqual(
            list(user.realms.all().values_list('workspace', 'partner', 'group__name', 'is_active')),
            [
                (user.workspace.id, user.partner.id, Group.objects.get(name='IP_VIEWER').name, True),
                (user.workspace.id, user.partner.id, Group.objects.get(name='IP_EDITOR').name, True),
            ]
        )

    def test_auth_required(self):
        self.client.force_authenticate()
        response = self.client.post(reverse('user-realms-import'), data={})
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_empty_realms(self):
        """
        If realms is an empty list, no user realms will be active and user is deactivated as well
        """
        user = PartnerUserFactory(realms__data=['IP_VIEWER'])
        input_data = {
            'email': user.email,
            'first_name': user.first_name,
            'middle_name': user.middle_name,
            'last_name': user.last_name,
            'realms': []
        }
        self.client.force_authenticate(factories.NonPartnerUserFactory())
        response = self.client.post(reverse('user-realms-import'), data=input_data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        user.refresh_from_db()
        self.assertFalse(user.is_active)
        self.assertFalse(user.realms.filter(is_active=True).exists())

    def test_issue_email(self):
        user = PartnerUserFactory.build(realms__data=[], email='new_user@domain-with-dash.org')
        workspace = WorkspaceFactory()
        partner = PartnerFactory()
        self.assertFalse(get_user_model().objects.filter(email=user.email).exists())
        input_data = {
            'email': user.email,
            'first_name': user.first_name,
            'middle_name': user.middle_name,
            'last_name': user.last_name,
            'realms': [
                {
                    'country': workspace.external_id,
                    'organization': partner.vendor_number,
                    'group': "IP Editor",
                },
                {
                    'country': workspace.external_id,
                    'organization': partner.vendor_number,
                    'group': "IP Viewer",
                },
                {
                    'country': workspace.external_id,
                    'organization': partner.vendor_number,
                    'group': "IP LM Editor",
                },
            ]
        }
        self.client.force_authenticate(factories.NonPartnerUserFactory())
        response = self.client.post(reverse('user-realms-import'), data=input_data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)

        self.assertTrue(get_user_model().objects.filter(email=user.email).exists())
        user = get_user_model().objects.get(email=user.email)
        self.assertCountEqual(
            list(user.realms.all().values_list('workspace', 'partner', 'group__name', 'is_active')),
            [
                (user.workspace.id, user.partner.id, Group.objects.get(name='IP_VIEWER').name, True),
                (user.workspace.id, user.partner.id, Group.objects.get(name='IP_EDITOR').name, True),
            ]
        )


class TestProgrammeDocumentCalculationMethodsAPIView(BaseProgressReportAPITestCase):
    def setUp(self):
        super().setUp()
        self.llo2 = factories.LowerLevelOutputFactory(cp_output=self.cp_output)

    def test_get_calculation_methods_all_active(self):
        url = reverse(
            'programme-document-calculation-methods',
            args=[self.workspace.pk, self.pd.pk],
        )
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['ll_outputs_and_indicators'].__len__(), self.pd.lower_level_outputs.count(), 2)
        self.assertEqual(
            sorted([llo['ll_output']['id'] for llo in response.data['ll_outputs_and_indicators']]),
            [self.llo.id, self.llo2.id])

    def test_get_calculation_methods_filter_inactive(self):
        self.llo2.active = False
        self.llo2.save()
        url = reverse(
            'programme-document-calculation-methods',
            args=[self.workspace.pk, self.pd.pk],
        )
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['ll_outputs_and_indicators'].__len__(), 1)
        self.assertEqual(response.data['ll_outputs_and_indicators'][0]['ll_output']['id'], self.llo.id)
