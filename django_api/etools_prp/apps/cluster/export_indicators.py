import itertools
import uuid

from django.conf import settings
from django.db.models import Count

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

from etools_prp.apps.indicator.models import DisaggregationValue, IndicatorBlueprint, IndicatorReport

PATH = settings.BASE_DIR + "/apps/cluster/templates/excel/indicators_export.xlsx"
SAVE_PATH = '/tmp/'

DISAGGREGATION_COLUMN_START = 44
INDICATOR_DATA_ROW_START = 5
MAXIMUM_DISAGGREGATIONS_PER_INDICATOR = 3


class IndicatorsXLSXExporter:

    analysis = False

    def __init__(self, indicators, response_plan_id, analysis=None):
        self.wb = load_workbook(PATH)
        self.sheet = self.wb.get_active_sheet()
        self.indicators = indicators
        self.response_plan_id = response_plan_id
        if analysis is not None:
            self.analysis = analysis
        self.sheets = [self.sheet, ]
        self.disaggregations_start_column = DISAGGREGATION_COLUMN_START

        self.bold_center_style = NamedStyle(name="Bold and Center")
        self.bold_center_style.font = Font(bold=True)
        self.bold_center_style.alignment = Alignment(horizontal='center')

    def duplicate_sheet(self, sheet):
        return self.wb.copy_worksheet(sheet)

    def fill_sheet(self, disaggregation_types, indicators):
        # Setup a title (limited to 30 chars via excel requirements)
        self.sheet.title = ", ".join(
            [dt.name for dt in disaggregation_types]
        )[:30] if disaggregation_types else "None"

        # Generate disaggregation types columns
        # it holds column number for given Disaggregation Type ID
        disaggregation_types_map = dict()
        for idx, dt in enumerate(disaggregation_types):
            current_column = self.disaggregations_start_column + idx

            name_cell = self.sheet.cell(row=1, column=current_column, value=dt.name)
            name_cell.alignment = Alignment(horizontal='center')
            name_cell.font = Font(bold=True)

            id_cell = self.sheet.cell(row=2, column=current_column, value=dt.id)
            id_cell.alignment = Alignment(horizontal='center')

            type_cell = self.sheet.cell(
                row=4, column=current_column, value="#indicator+type+" + dt.name
            )
            type_cell.alignment = Alignment(horizontal='center')

            disaggregation_types_map[dt.id] = current_column

        # Prepare Disaggregation Values for given disaggregation types
        disaggregation_values = list()
        disaggregation_values_list = list()
        disaggregation_values_map = dict()
        disaggregation_values_by_type = dict()
        for disaggregation_type in disaggregation_types:
            disaggregation_values_base = DisaggregationValue.objects.filter(
                disaggregation=disaggregation_type
            ).order_by('value')
            disaggregation_values_list.append(disaggregation_values_base)
            for dv in disaggregation_values_base:
                disaggregation_values_by_type[dv.id] = disaggregation_type.name

        # Create all possible combinations (max. 3 items per combination) for
        # disaggregation values
        for combination_items in range(1, MAXIMUM_DISAGGREGATIONS_PER_INDICATOR + 1):
            for combinations in itertools.combinations(
                    disaggregation_values_list, combination_items):
                disaggregation_values += list(itertools.product(*combinations))

        # Generate disaggregation values columns
        for idx, dvs in enumerate(disaggregation_values):
            current_column = self.disaggregations_start_column + len(disaggregation_types) + idx

            cell = self.sheet.cell(row=1, column=current_column, value=" + ".join([dv.value for dv in dvs]))
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)

            self.sheet.cell(row=2, column=current_column, value=", ".join([str(dv.id) for dv in dvs]))
            self.sheet.cell(
                row=3, column=current_column, value=" + ".join([disaggregation_values_by_type[dv.id] for dv in dvs])
            )
            self.sheet.cell(
                row=4, column=current_column, value="#indicator+value+" + str("+".join([dv.value for dv in dvs]))
            )

            disaggregation_values_map[", ".join([str(dv.id) for dv in dvs])] = current_column

        totals_column = self.disaggregations_start_column + len(disaggregation_types) + len(disaggregation_values)
        totals_header_cell = self.sheet.cell(row=1, column=totals_column, value="Total")
        totals_header_cell.alignment = Alignment(horizontal='center')

        disaggregation_values_map["()"] = totals_column

        # Each row is one location per indicator
        start_row_id = INDICATOR_DATA_ROW_START

        all_indicators = list()
        # Dual Reporting
        for indicator in indicators:
            all_indicators.append(indicator)
            for children_indicator in indicator.children.all():
                all_indicators.append(children_indicator)

        for indicator in all_indicators:
            cluster_objective = indicator.reportable.cluster_objectives.first()
            cluster_activity = indicator.reportable.cluster_activities.first()
            partner_project = indicator.reportable.partner_projects.first()
            partner_activity = None
            partner_activity_project_context = indicator.reportable.partner_activity_project_contexts.first()
            locations_data = indicator.indicator_location_data

            cluster = None

            if cluster_objective:
                cluster = cluster_objective.cluster
                partner_activity = cluster_objective.cluster_activities.first().partner_activities.first() \
                    if cluster_objective.cluster_activities.first() else ""
                partner_project = cluster_objective.cluster.partner_projects.first()
            elif cluster_activity:
                cluster = cluster_activity.cluster_objective.cluster
                cluster_objective = cluster_activity.cluster_objective
                partner_activity = cluster_activity.partner_activities.first()
                partner_project = cluster_activity.cluster_objective.cluster.partner_projects.first()
            elif partner_activity_project_context:
                partner_activity = partner_activity_project_context.activity
                cluster = partner_activity.cluster_activity.cluster_objective.cluster if \
                    partner_activity.cluster_activity else partner_activity.cluster_objective.cluster
                cluster_objective = partner_activity.cluster_activity.cluster_objective if \
                    partner_activity.cluster_activity else partner_activity.cluster_objective
                partner_project = partner_activity_project_context.project
            elif partner_project:
                cluster = partner_project.clusters.first()
                cluster_objective = cluster.cluster_objectives.first()
                partner_activity = partner_project.partner.partner_activities.first()

            for location_data in locations_data.all():

                self.sheet.cell(
                    row=start_row_id,
                    column=1).value = location_data.location.gateway.country.name
                self.sheet.cell(
                    row=start_row_id,
                    column=2).value = location_data.location.gateway.country.country_short_code
                self.sheet.cell(row=start_row_id,
                                column=3).value = cluster.response_plan.title if cluster else ""

                self.sheet.cell(row=start_row_id,
                                column=4).value = cluster.get_type_display() if cluster else ""
                self.sheet.cell(
                    row=start_row_id,
                    column=5).value = (cluster.partner_projects.first().partner.title if
                                       cluster.partner_projects.first() else "") if cluster else ""
                self.sheet.cell(
                    row=start_row_id,
                    column=6).value = cluster_objective.title if cluster_objective else ""
                self.sheet.cell(
                    row=start_row_id,
                    column=7).value = partner_activity.title if partner_activity else ""
                self.sheet.cell(
                    row=start_row_id,
                    column=8).value = indicator.reportable.blueprint.title
                self.sheet.cell(
                    row=start_row_id,
                    column=9).value = indicator.reportable.blueprint.display_type
                self.sheet.cell(
                    row=start_row_id,
                    column=10).value = indicator.reportable.blueprint.get_calculation_formula_across_locations_display()
                self.sheet.cell(
                    row=start_row_id,
                    column=11).value = indicator.reportable.blueprint.get_calculation_formula_across_periods_display()

                self.sheet.cell(
                    row=start_row_id,
                    column=12).value = partner_project.title if partner_project else ""
                self.sheet.cell(
                    row=start_row_id,
                    column=13).value = partner_project.get_status_display() if partner_project else ""
                self.sheet.cell(row=start_row_id,
                                column=14).value = partner_activity_project_context.start_date if partner_activity_project_context else ""
                self.sheet.cell(row=start_row_id,
                                column=15).value = partner_activity_project_context.end_date if partner_activity_project_context else ""

                self.sheet.cell(
                    row=start_row_id,
                    column=16).value = indicator.reportable.calculated_baseline

                self.sheet.cell(row=start_row_id,
                                column=17).value = indicator.reportable.calculated_target
                self.sheet.cell(
                    row=start_row_id,
                    column=18).value = indicator.reportable.total['c']

                self.sheet.cell(row=start_row_id,
                                column=19).value = location_data.location.title
                self.sheet.cell(
                    row=start_row_id,
                    column=20).value = location_data.location.gateway.name

                # Iterate over location admin references:
                location = location_data.location
                while True:
                    admin_level = location.gateway.admin_level
                    # TODO: secure in case of wrong location data
                    admin_level = min(admin_level, 5)
                    self.sheet.cell(row=start_row_id, column=20 + admin_level * 2).value = location.title
                    self.sheet.cell(row=start_row_id, column=20 + admin_level * 2 - 1).value = location.gateway.name

                    if location.parent:
                        location = location.parent
                    else:
                        break

                self.sheet.cell(row=start_row_id,
                                column=31).value = indicator.time_period_start
                self.sheet.cell(row=start_row_id,
                                column=32).value = indicator.time_period_end

                self.sheet.cell(
                    row=start_row_id,
                    column=33).value = indicator.get_report_status_display()
                self.sheet.cell(
                    row=start_row_id,
                    column=34).value = indicator.get_overall_status_display()

                self.sheet.cell(row=start_row_id,
                                column=35).value = indicator.submission_date
                self.sheet.cell(row=start_row_id, column=36).value = cluster.id if cluster else ""
                self.sheet.cell(
                    row=start_row_id,
                    column=37).value = cluster_objective.id if cluster_objective else ""
                self.sheet.cell(
                    row=start_row_id,
                    column=38).value = partner_activity.id if partner_activity else ""
                self.sheet.cell(
                    row=start_row_id,
                    column=39).value = indicator.reportable.blueprint.id
                self.sheet.cell(row=start_row_id,
                                column=40).value = partner_project.partner.id if partner_project else \
                    (partner_activity.partner.id if partner_activity else "")
                self.sheet.cell(
                    row=start_row_id,
                    column=41).value = partner_project.id if partner_project else ""
                self.sheet.cell(
                    row=start_row_id,
                    column=42).value = indicator.id
                self.sheet.cell(
                    row=start_row_id,
                    column=43).value = location_data.id

                # Check location item disaggregation type
                for reported_disaggregation_type in location_data.disaggregation_reported_on:
                    if reported_disaggregation_type in disaggregation_types_map:
                        self.sheet.cell(
                            row=start_row_id,
                            column=disaggregation_types_map[reported_disaggregation_type]).value = "X"

                # Check location item values
                blueprint = location_data.indicator_report.reportable.blueprint

                for k, v in location_data.disaggregation.items():
                    if k == "()":
                        self.sheet.cell(
                            row=start_row_id,
                            column=disaggregation_values_map['()']).value = "" if v['d'] == 0 else \
                            (v['v'] if blueprint.unit == IndicatorBlueprint.NUMBER else "{}/{}".format(v['v'], v['d']))

                    else:
                        for dk, dv in disaggregation_values_map.items():
                            if dk == "()":
                                continue
                            if sorted(list(eval(k))) == sorted(
                                    list(int(k) for k in dk.split(","))):
                                self.sheet.cell(
                                    row=start_row_id, column=dv).value = v['v'] \
                                        if blueprint.unit == IndicatorBlueprint.NUMBER \
                                        else "{}/{}".format(v['v'], v['d'])

                start_row_id += 1

        # Lock first rows
        self.sheet.freeze_panes = 'A%d' % INDICATOR_DATA_ROW_START

        return True

    def merge_sheets(self):
        """
        1. Create new sheet for Analysis
        2. Copy generic columns (same for all)
        3. Merge disaggregation types columns
        4. Merge disaggregation values columns
        5. Add Total column at the very end of sheet
        6. Iterate over basic sheets and merge data into Analysis sheet
        7. Remove basic sheets
        """

        # Create new sheet
        merged_sheet = self.wb.create_sheet(title="Analysis")

        # Copy generic columns
        for i in range(1, self.disaggregations_start_column):
            merged_sheet.cell(
                column=i, row=1, value=self.sheet.cell(column=i, row=1).value
            ).style = self.bold_center_style

            merged_sheet.cell(
                column=i, row=2, value=self.sheet.cell(column=i, row=2).value
            ).alignment = Alignment(horizontal='center')

            merged_sheet.cell(
                column=i, row=3, value=self.sheet.cell(column=i, row=3).value
            ).alignment = Alignment(horizontal='center')

            merged_sheet.cell(
                column=i, row=4, value=self.sheet.cell(column=i, row=4).value
            ).alignment = Alignment(horizontal='center')

        # Merge disaggregation types
        merged_disaggregations = list()
        merged_disaggregations_map = dict()

        for sheet in self.sheets:
            column = self.disaggregations_start_column
            while True:
                if sheet.cell(column=column, row=4).value is None or sheet.cell(
                        column=column, row=4).value.find("#indicator+type+") < 0:
                    break
                if (
                    sheet.cell(column=column, row=1).value,
                    sheet.cell(column=column, row=2).value,
                    sheet.cell(column=column, row=4).value
                ) not in merged_disaggregations:
                    merged_disaggregations.append(
                        (
                            sheet.cell(column=column, row=1).value,
                            sheet.cell(column=column, row=2).value,
                            sheet.cell(column=column, row=4).value
                        )
                    )
                column += 1

        # Prepare new column headers for disaggregation types
        for idx, (
            disaggregation_name,
            disaggregation_id,
            disaggregation_label,
        ) in enumerate(merged_disaggregations):
            current_column = self.disaggregations_start_column + idx
            name_cell = merged_sheet.cell(column=current_column, row=1, value=disaggregation_name)
            name_cell.style = self.bold_center_style

            id_cell = merged_sheet.cell(column=current_column, row=2, value=disaggregation_id)
            id_cell.alignment = Alignment(horizontal='center')

            label_cell = merged_sheet.cell(column=current_column, row=4, value=disaggregation_label)
            label_cell.alignment = Alignment(horizontal='center')

            merged_disaggregations_map[disaggregation_id] = current_column

        # Merge disaggregation values
        merged_disaggregation_values = list()
        merged_disaggregation_values_map = dict()
        for sheet in self.sheets:
            column = self.disaggregations_start_column
            while True:
                if sheet.cell(column=column, row=4).value and sheet.cell(
                        column=column, row=4).value.find("#indicator+type+") > -1:
                    column += 1
                    continue
                if sheet.cell(column=column, row=4).value is None or sheet.cell(
                        column=column, row=4).value.find("#indicator+value+") < 0:
                    break
                if (
                        sheet.cell(column=column, row=1).value,
                        sheet.cell(column=column, row=2).value,
                        sheet.cell(column=column, row=3).value,
                        sheet.cell(column=column, row=4).value
                ) not in merged_disaggregation_values:
                    merged_disaggregation_values.append(
                        (
                            sheet.cell(column=column, row=1).value,
                            sheet.cell(column=column, row=2).value,
                            sheet.cell(column=column, row=3).value,
                            sheet.cell(column=column, row=4).value
                        )
                    )
                column += 1

        # Prepare new column headers for disaggregation values
        for idx, (
            disaggregation_value,
            disaggregation_id,
            disaggregation_name,
            disaggregation_label,
        ) in enumerate(merged_disaggregation_values):
            current_column = self.disaggregations_start_column + len(merged_disaggregations) + idx
            value_cell = merged_sheet.cell(column=current_column, row=1, value=disaggregation_value)
            value_cell.style = self.bold_center_style

            id_cell = merged_sheet.cell(column=current_column, row=2, value=disaggregation_id)
            id_cell.alignment = Alignment(horizontal='center')

            name_cell = merged_sheet.cell(column=current_column, row=3, value=disaggregation_name)
            name_cell.alignment = Alignment(horizontal='center')

            label_cell = merged_sheet.cell(column=current_column, row=4, value=disaggregation_label)
            label_cell.alignment = Alignment(horizontal='center')

            merged_disaggregation_values_map[disaggregation_id] = current_column

        # Add Total column at the very end
        totals_column = self.disaggregations_start_column + \
            len(merged_disaggregations) + \
            len(merged_disaggregation_values)

        totals_header_cell = merged_sheet.cell(column=totals_column, row=1, value="Total")
        totals_header_cell.style = self.bold_center_style

        # Merge data from all sheets
        merged_row = INDICATOR_DATA_ROW_START
        for sheet in self.sheets:
            # Count columns
            max_columns = 1
            while True:
                # Stop condition
                if sheet.cell(column=max_columns, row=1).value is None:
                    max_columns -= 1
                    break
                max_columns += 1
            sheet_row = INDICATOR_DATA_ROW_START
            while True:
                # Stop row condition
                if sheet.cell(column=1, row=sheet_row).value is None:
                    break
                for column in range(max_columns):
                    column += 1
                    # If value less then self.disaggregations_start_column just copy
                    # data
                    if column < self.disaggregations_start_column:
                        merged_sheet.cell(
                            column=column, row=merged_row).value = sheet.cell(
                            column=column, row=sheet_row).value
                    else:
                        # Total has None value as a type
                        if sheet.cell(column=column, row=4).value is None:
                            col = self.disaggregations_start_column + \
                                  len(merged_disaggregations) + \
                                  len(merged_disaggregation_values)
                            merged_sheet.cell(
                                column=col, row=merged_row
                            ).value = sheet.cell(column=column, row=sheet_row).value
                        # If column is disaggregation type
                        elif sheet.cell(column=column, row=4).value.find("#indicator+type+") > -1:
                            merged_sheet.cell(
                                column=merged_disaggregations_map[sheet.cell(column=column, row=2).value],
                                row=merged_row
                            ).value = sheet.cell(column=column, row=sheet_row).value
                        # If column is disaggregation value
                        elif sheet.cell(column=column, row=4).value.find("#indicator+value+") > -1:
                            merged_sheet.cell(
                                column=merged_disaggregation_values_map[sheet.cell(column=column, row=2).value],
                                row=merged_row
                            ).value = sheet.cell(column=column, row=sheet_row).value

                merged_row += 1
                sheet_row += 1
            # Remove basic sheets
            self.wb.remove_sheet(sheet)

        # Adjust columns width
        for col in merged_sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except BaseException:
                    pass
            adjusted_width = max_length + 2
            merged_sheet.column_dimensions[column].width = adjusted_width

        # Add filters
        merged_sheet.auto_filter.ref = "A1:%s%d" % (
            get_column_letter(
                self.disaggregations_start_column +
                len(merged_disaggregations) +
                len(merged_disaggregation_values)),
            merged_row
        )

        # Lock first rows
        merged_sheet.freeze_panes = 'A%d' % INDICATOR_DATA_ROW_START

    def export_data(self):

        # Prepare list of unique disaggregations for choosed indicators
        disaggregation_types_base_set = set()
        for i in self.indicators:
            for d in i.reportable.disaggregations.all():
                disaggregation_types_base_set.add(d)
        disaggregation_types_base = list(disaggregation_types_base_set)

        # Split data into spreadsheets, combination of
        # disaggregation_types_length with r-length tuples and no repeated
        # elements
        disaggregation_types_list = list()
        for i in range(MAXIMUM_DISAGGREGATIONS_PER_INDICATOR + 1):
            disaggregation_types_list += list(
                itertools.combinations(
                    disaggregation_types_base, i))

        # Generate data per spreadsheet
        sheet_no = 1
        to_remove = list()
        for disaggregation_types in disaggregation_types_list:
            # Check if indicators belongs to this disaggregation_types
            # Skip if not
            indicators = self.indicators
            indicators = IndicatorReport.objects.filter(pk__in=[i.id for i in indicators])
            if disaggregation_types:
                indicators = indicators.annotate(
                    count=Count('reportable__disaggregations')).filter(
                    count=len(disaggregation_types))
                for dt in disaggregation_types:
                    indicators = indicators.filter(reportable__disaggregations=dt)
            else:
                indicators = indicators.filter(reportable__disaggregations__isnull=True)

            indicators = indicators.order_by('reportable__id')
            if not indicators:
                continue

            self.sheets.append(self.duplicate_sheet(self.sheets[0]))
            self.sheet = self.sheets[sheet_no]

            if not self.fill_sheet(disaggregation_types, indicators):
                to_remove.append(self.sheets[sheet_no])
            sheet_no += 1

        to_remove.append(self.sheets[0])

        # Remove empty spreadsheets
        for s in to_remove:
            # Spreadsheet need atleast 1 sheet
            if len(self.sheets) > 1:
                self.sheets.remove(s)
                self.wb.remove_sheet(s)

        if self.analysis:
            self.merge_sheets()

        file_path = SAVE_PATH + 'export_' + uuid.uuid4().hex + '.xlsx'
        self.wb.save(file_path)
        return file_path
