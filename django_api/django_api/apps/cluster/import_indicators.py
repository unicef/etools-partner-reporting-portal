# -*- coding: utf-8 -*-
import traceback
from datetime import datetime
from django.db.models import Q
from django.db import transaction

from openpyxl.reader.excel import load_workbook

from indicator.models import IndicatorLocationData, IndicatorBlueprint, DisaggregationValue
from indicator.disaggregators import QuantityIndicatorDisaggregator, RatioIndicatorDisaggregator

COLUMN_HASH_ID = 4
MAX_COLUMNS = 1000


class IndicatorsXLSXReader(object):

    def __init__(self, path, partner):
        self.wb = load_workbook(path)
        self.partner = partner

    def import_data(self):
        for self.sheet in self.wb.worksheets:
            # Find "Location ID" column
            location_column_id = None
            for column in range(1, MAX_COLUMNS):
                if self.sheet.cell(row=COLUMN_HASH_ID, column=column).value == "#loc+id":
                    location_column_id = column
                    break
            if not location_column_id:
                return "Cannot find Location ID column"

            # Find "Total" column
            total_column_id = None
            for column in range(1, MAX_COLUMNS):
                if self.sheet.cell(row=1, column=column).value == "Total":
                    total_column_id = column
                    break
            if not total_column_id:
                return "Cannot find Total column"

            # Find first Disaggregation Value column
            dis_data_column_start_id = None
            for column in range(1, MAX_COLUMNS):
                if not self.sheet.cell(row=COLUMN_HASH_ID, column=column).value:
                    break
                if "#indicator+value" in self.sheet.cell(row=COLUMN_HASH_ID, column=column).value:
                    dis_data_column_start_id = column
                    break

            # Iterate over rows and save disaggregation values
            for row in range(COLUMN_HASH_ID + 1, self.sheet.max_row):
                # If row is empty, end of sheet
                if not self.sheet.cell(row=row, column=1).value:
                    break

                # Get IndicatorLocationData ID
                try:
                    ild_id = str(int(self.sheet.cell(row=row, column=location_column_id).value))

                    ind = IndicatorLocationData.objects.filter(pk=ild_id)

                    # Check if indicator has parent (UNICEF)
                    # If does, use parent to check partner
                    if self.partner and ind.filter(indicator_report__parent__isnull=False):
                        if not ind.filter(
                                indicator_report__parent__reportable__partner_activity_project_contexts__project__partner=self.partner
                        ).exists():
                            return "Parent of Indicator ID " \
                                + ild_id \
                                + " does not belong to partner " \
                                + str(self.partner)
                    # Check if Partner is allowed to modify data
                    elif self.partner and ind.filter(
                            Q(**{
                                'indicator_report__reportable__cluster_objectives'
                                '__cluster__partner_projects__partner': self.partner
                            }) |
                            Q(**{
                                'indicator_report__reportable'
                                '__cluster_objectives__cluster_activities__partner_activities'
                                '__partner': self.partner
                            }) |
                            Q(**{
                                'indicator_report__reportable'
                                '__cluster_activities__cluster_objective__cluster__partner_projects'
                                '__partner': self.partner
                            }) |
                            Q(**{
                                'indicator_report__reportable'
                                '__cluster_activities__partner_activities'
                                '__partner': self.partner
                            }) |
                            Q(**{
                                'indicator_report__reportable'
                                '__partner_activity_project_contexts__project'
                                '__partner': self.partner
                            }) |
                            Q(**{
                                'indicator_report__reportable'
                                '__partner_projects'
                                '__partner': self.partner
                            })).count() == 0:
                        return "Indicator ID " + ild_id + " does not belong to partner " + str(self.partner)
                    indicator = IndicatorLocationData.objects.get(
                        pk=int(self.sheet.cell(row=row, column=location_column_id).value)
                    )

                    # Check if Indicator Report is not able to submit anymore
                    if not indicator.indicator_report.can_import:
                        transaction.rollback()
                        return "Indicator in row {} is already submitted. Please remove row and try again.".format(
                            row,)

                except IndicatorLocationData.DoesNotExist:
                    return "Cannot find Indicator Location Data data for ID " \
                        + str(self.sheet.cell(row=row, column=location_column_id).value)

                blueprint = indicator.indicator_report.reportable.blueprint
                data = indicator.disaggregation
                # Prepare
                already_updated_row_value = False
                for column in range(dis_data_column_start_id if dis_data_column_start_id else total_column_id,
                                    total_column_id + 1):
                    try:
                        value = self.sheet.cell(row=row, column=column).value
                        # Check if value is present in cell
                        if value is not None:
                            # Evaluate ID of Disaggregation Type
                            dis_type_id = "()"
                            dis_type_value = self.sheet.cell(row=2, column=column).value
                            if dis_type_value:
                                dis_type_value = sorted(list(map(int, str(dis_type_value).split(","))), key=int)
                                dis_type_id = str(tuple(dis_type_value))

                            if dis_type_id not in data:
                                # Check if data is proper disaggregation value
                                for dt in dis_type_value:
                                    if not DisaggregationValue.objects.filter(pk=dt).exists():
                                        transaction.rollback()
                                        return "Disaggregation {} does not exists".format(
                                            self.sheet.cell(row=4, column=column).value)
                                    # Check if filled disaggregation values belongs to their type
                                    dv = DisaggregationValue.objects.get(pk=dt)
                                    if dv.disaggregation.id not in indicator.disaggregation_reported_on:
                                        transaction.rollback()
                                        return "Disaggregation {} does not belong to this Indicator".format(
                                            self.sheet.cell(row=4, column=column).value)
                                # Create value
                                data[dis_type_id] = dict()

                            already_updated_row_value = True

                            # Update values
                            if blueprint.unit == IndicatorBlueprint.NUMBER:
                                data[dis_type_id]["v"] = value
                            else:
                                if isinstance(value, datetime):
                                    transaction.rollback()
                                    return "Value in column {}, row {} is Date Time. Please format row to Plain Text."\
                                        .format(self.sheet.cell(row=4, column=column).value, row)
                                values = value.split("/")
                                data[dis_type_id]["v"] = int(values[0])
                                data[dis_type_id]["d"] = int(values[1])
                        else:
                            # if value is not present, check if it should be
                            # all rows need to updated

                            # Evaluate ID of Disaggregation Type
                            dis_type_value = self.sheet.cell(row=2, column=column).value
                            if dis_type_value:
                                dis_type_value = sorted(list(map(int, str(dis_type_value).split(","))), key=int)
                            if dis_type_value:
                                if len(dis_type_value) == indicator.level_reported:
                                    # Check if data is proper disaggregation value
                                    for dt in dis_type_value:
                                        dv = DisaggregationValue.objects.get(pk=dt)
                                        if dv.disaggregation.id in indicator.disaggregation_reported_on and \
                                                already_updated_row_value:
                                            transaction.rollback()
                                            return "Please fulfill required value to column {}, row {}"\
                                                .format(self.sheet.cell(row=4, column=column).value, row)

                    except Exception:
                        traceback.print_exc()
                        transaction.rollback()
                        return "Cannot assign disaggregation value to column {}, row {}"\
                            .format(self.sheet.cell(row=4, column=column).value, row)

                indicator.disaggregation = data
                indicator.save()

                if blueprint.unit == IndicatorBlueprint.NUMBER:
                    QuantityIndicatorDisaggregator.post_process(indicator)

                if blueprint.unit == IndicatorBlueprint.PERCENTAGE:
                    RatioIndicatorDisaggregator.post_process(indicator)

        return
