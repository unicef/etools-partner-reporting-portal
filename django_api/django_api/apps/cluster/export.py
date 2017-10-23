from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.db.models import Count

import itertools

from indicator.models import Disaggregation, DisaggregationValue

PATH = settings.BASE_DIR + "/apps/cluster/templates/excel/export.xlsx"
SAVE_PATH = settings.MEDIA_ROOT + '/'

DISAGGREGATION_COLUMN_START = 44
INDICATOR_DATA_ROW_START = 5
MAXIMUM_DISSAGREGATIONS_PER_INDICATOR = 3

class XLSXWriter:
    def __init__(self, indicators, response_plan_id, analysis=False):

        self.wb = load_workbook(PATH)
        self.sheet = self.wb.get_active_sheet()
        self.indicators = indicators
        self.response_plan_id = response_plan_id
        self.analysis = analysis


    def duplicate_sheet(self, sheet):
        return self.wb.copy_worksheet(sheet)


    def generate_sheet(self, disaggregation_types):

        # Setup a title
        self.sheet.title = ",".join([dt.name for dt in disaggregation_types]) if disaggregation_types else "None"

        # Generate disaggregation types columns
        disaggregation_types_map = dict() # it holds column number for given Disaggregation Type ID
        for idx, dt in enumerate(disaggregation_types):
            # Name
            self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + idx).value = dt.name
            self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + idx).alignment = Alignment(horizontal='center')
            self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + idx).font = Font(bold=True)
            # ID
            self.sheet.cell(row=2, column=DISAGGREGATION_COLUMN_START + idx).value = dt.id
            self.sheet.cell(row=2, column=DISAGGREGATION_COLUMN_START + idx).alignment = Alignment(horizontal='center')
            # Type
            self.sheet.cell(row=4, column=DISAGGREGATION_COLUMN_START + idx).value = "#indicator+type+" + dt.name
            self.sheet.cell(row=4, column=DISAGGREGATION_COLUMN_START + idx).alignment = Alignment(horizontal='center')
            disaggregation_types_map[dt.id] = DISAGGREGATION_COLUMN_START + idx


        # Prepare Disaggregation Values for given disaggregation types
        disaggregation_values = list()
        disaggregation_values_list = list()
        disaggregation_values_map = dict()
        disaggregation_values_by_type = dict()
        for disaggregation_type in disaggregation_types:
                disaggregation_values_base = DisaggregationValue.objects.filter(disaggregation=disaggregation_type).order_by('value')
                disaggregation_values_list.append(disaggregation_values_base)
                for dv in disaggregation_values_base:
                    disaggregation_values_by_type[dv.id] = disaggregation_type.name

        # Create all possible combinations (max. 3 items per combination) for disaggregation values
        for combination_items in range(1, 4):
            for combinations in itertools.combinations(disaggregation_values_list, combination_items):
                disaggregation_values += list(itertools.product(*combinations))

        # Generate disaggregation values columns
        for idx, dvs in enumerate(disaggregation_values):
            self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + idx).value =  " + ".join([dv.value for dv in dvs])
            self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + idx).alignment = Alignment(horizontal='center')
            self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + idx).font = Font(bold=True)
            self.sheet.cell(row=2, column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + idx).value = ", ".join([str(dv.id) for dv in dvs])
            self.sheet.cell(row=3, column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + idx).value = " + ".join([disaggregation_values_by_type[dv.id] for dv in dvs])
            self.sheet.cell(row=4, column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + idx).value = "#indicator+value+" + str("+".join([dv.value for dv in dvs]))
            disaggregation_values_map[", ".join([str(dv.id) for dv in dvs])] = DISAGGREGATION_COLUMN_START + len(disaggregation_types) + idx

        # Generate Total disaggregation value
        self.sheet.cell(row=1,column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + len(disaggregation_values)).value = "Total"
        self.sheet.cell(row=1, column=DISAGGREGATION_COLUMN_START + len(disaggregation_types) + len(disaggregation_values)).alignment = Alignment(horizontal='center')
        disaggregation_values_map["()"] = DISAGGREGATION_COLUMN_START + len(disaggregation_types) + len(disaggregation_values)

        indicators = self.indicators
        if disaggregation_types:
            indicators = indicators.annotate(count=Count('reportable__disaggregations')).filter(count=len(disaggregation_types))
            for dt in disaggregation_types:
                indicators = indicators.filter(reportable__disaggregations=dt)
        else:
            indicators = indicators.filter(reportable__disaggregations__isnull=True)

        indicators = indicators.order_by('reportable__id')

        if not indicators:
            return False

        # Each row is one location per indicator
        start_row_id = INDICATOR_DATA_ROW_START

        for indicator in indicators:
            cluster_objective = indicator.reportable.cluster_objectives.first()
            cluster_activity = indicator.reportable.cluster_activities.first()
            partner_project = indicator.reportable.partner_projects.first()
            partner_activity = indicator.reportable.partner_activities.first()
            locations_data = indicator.indicator_location_data

            cluster = None

            if cluster_objective:
                cluster = cluster_objective.cluster
                partner_activity = cluster_objective.cluster_activities.first().partner_activities.first()
                partner_project = cluster_objective.cluster.partner_projects.first()
            elif cluster_activity:
                cluster = cluster_activity.cluster_objective.cluster
                cluster_objective = cluster_activity.cluster_objective
                partner_activity = cluster_activity.partner_activities.first()
                partner_project = cluster_activity.cluster_objective.cluster.partner_projects.first()
            elif partner_activity:
                cluster = partner_activity.cluster_activity.cluster_objective.cluster
                cluster_objective = partner_activity.cluster_activity.cluster_objective
                partner_project = partner_activity.project
            elif partner_project:
                cluster = partner_project.clusters.first()
                cluster_objective = partner_project.clusters.first().cluster_objectives.first()
                partner_activity = partner_project.partner.partner_activities.first()

            for location_data in locations_data.all():

                self.sheet.cell(row=start_row_id, column=1).value = location_data.location.gateway.country.name
                self.sheet.cell(row=start_row_id, column=2).value = location_data.location.gateway.country.country_short_code
                self.sheet.cell(row=start_row_id, column=3).value = cluster.response_plan.title

                self.sheet.cell(row=start_row_id, column=4).value = cluster.get_type_display()
                self.sheet.cell(row=start_row_id, column=5).value = cluster.partner_projects.first().partner.title
                self.sheet.cell(row=start_row_id, column=6).value = cluster_objective.title
                self.sheet.cell(row=start_row_id, column=7).value = partner_activity.title
                self.sheet.cell(row=start_row_id, column=8).value = indicator.reportable.blueprint.title

                self.sheet.cell(row=start_row_id, column=9).value = indicator.reportable.blueprint.get_unit_display()
                self.sheet.cell(row=start_row_id, column=10).value = indicator.reportable.blueprint.get_calculation_formula_across_locations_display()
                self.sheet.cell(row=start_row_id, column=11).value = indicator.reportable.blueprint.get_calculation_formula_across_periods_display()

                self.sheet.cell(row=start_row_id, column=12).value = partner_project.title
                self.sheet.cell(row=start_row_id, column=13).value = partner_project.get_status_display()
                self.sheet.cell(row=start_row_id, column=14).value = partner_activity.start_date
                self.sheet.cell(row=start_row_id, column=15).value = partner_activity.end_date

                self.sheet.cell(row=start_row_id, column=16).value = indicator.reportable.baseline
                self.sheet.cell(row=start_row_id, column=17).value = indicator.reportable.target
                self.sheet.cell(row=start_row_id, column=18).value = indicator.reportable.total['c']

                self.sheet.cell(row=start_row_id, column=19).value = location_data.location.title
                self.sheet.cell(row=start_row_id, column=20).value = location_data.location.gateway.name

                # Iterate over location admin references:
                location = location_data.location
                while(True):
                    level = location.gateway.admin_level
                    # TODO: secure in case of wrong location data
                    if level > 5:
                        level = 5
                    self.sheet.cell(row=start_row_id, column=20 + level * 2).value = location.p_code
                    self.sheet.cell(row=start_row_id, column=20 + level * 2 - 1).value = location.gateway.name
                    if location.parent:
                        location = location.parent
                    else:
                        break



                self.sheet.cell(row=start_row_id, column=31).value = indicator.time_period_start
                self.sheet.cell(row=start_row_id, column=32).value = indicator.time_period_end

                self.sheet.cell(row=start_row_id, column=33).value = indicator.get_report_status_display()
                self.sheet.cell(row=start_row_id, column=34).value = indicator.get_overall_status_display()

                self.sheet.cell(row=start_row_id, column=35).value = indicator.submission_date
                self.sheet.cell(row=start_row_id, column=36).value = cluster.id
                self.sheet.cell(row=start_row_id, column=37).value = cluster_objective.id
                self.sheet.cell(row=start_row_id, column=38).value = partner_activity.id
                self.sheet.cell(row=start_row_id, column=39).value = indicator.reportable.blueprint.id
                self.sheet.cell(row=start_row_id, column=40).value = partner_project.partner.id
                self.sheet.cell(row=start_row_id, column=41).value = partner_project.id
                self.sheet.cell(row=start_row_id, column=42).value = indicator.id
                self.sheet.cell(row=start_row_id, column=43).value = location_data.id

                # Check location item disaggregation type
                for reported_disaggregation_type in location_data.disaggregation_reported_on:
                    if reported_disaggregation_type in disaggregation_types_map:
                        self.sheet.cell(row=start_row_id, column=disaggregation_types_map[reported_disaggregation_type]).value = "X"

                # Check location item values
                for k, v in location_data.disaggregation.items():
                    if k == "()":
                        self.sheet.cell(row=start_row_id, column=disaggregation_values_map['()']).value = v['c']
                    else:
                        for dk, dv in disaggregation_values_map.items():
                            if dk == "()":
                                continue
                            if sorted(list(eval(k))) == sorted(list(int(k) for k in dk.split(","))):
                                self.sheet.cell(row=start_row_id, column=dv).value = v['c']

                start_row_id += 1

        # Lock first rows
        self.sheet.freeze_panes = 'A%d' % INDICATOR_DATA_ROW_START

        return True

    def merge_sheets(self):
        """
        1. Create new sheet for Analysis
        2. Copy generic columns (same for all)
        3. Merge disaggregation types columns
        4. Merge disaggregation values columns
        5. Add Total column at the very end of sheet
        6. Iterate over basic sheets and merge data into Analysis sheet
        7. Remove basic sheets
        """

        # Create new sheet
        merged_sheet = self.wb.create_sheet(title="Analysis")

        # Copy generic columns
        for i in range(1, DISAGGREGATION_COLUMN_START):
            merged_sheet.cell(column=i, row=1).value = self.sheet.cell(column=i, row=1).value
            merged_sheet.cell(column=i, row=1).font = Font(bold=True)
            merged_sheet.cell(column=i, row=1).alignment = Alignment(horizontal='center')
            merged_sheet.cell(column=i, row=2).value = self.sheet.cell(column=i, row=2).value
            merged_sheet.cell(column=i, row=2).alignment = Alignment(horizontal='center')
            merged_sheet.cell(column=i, row=3).value = self.sheet.cell(column=i, row=3).value
            merged_sheet.cell(column=i, row=3).alignment = Alignment(horizontal='center')
            merged_sheet.cell(column=i, row=4).value = self.sheet.cell(column=i, row=4).value
            merged_sheet.cell(column=i, row=4).alignment = Alignment(horizontal='center')

        # Merge disaggregation types
        merged_disaggregations = list()
        merged_disaggregations_map = dict()
        for sheet in self.sheets:
            column = DISAGGREGATION_COLUMN_START
            while(True):
                if sheet.cell(column=column, row=4).value == None or sheet.cell(column=column, row=4).value.find("#indicator+type+") < 0:
                    break
                if (
                        sheet.cell(column=column, row=1).value,
                        sheet.cell(column=column, row=2).value,
                        sheet.cell(column=column, row=4).value
                    ) not in merged_disaggregations:
                    merged_disaggregations.append(
                        (
                            sheet.cell(column=column, row=1).value,
                            sheet.cell(column=column, row=2).value,
                            sheet.cell(column=column, row=4).value
                        )
                    )
                column += 1

        # Prepare new column headers for disaggregation types
        for idx, merged_disaggregation in enumerate(merged_disaggregations):
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + idx, row=1).value = merged_disaggregation[0]
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + idx, row=1).font = Font(bold=True)
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + idx, row=1).alignment = Alignment(horizontal='center')
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + idx, row=2).value = merged_disaggregation[1]
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + idx, row=2).alignment = Alignment(horizontal='center')
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + idx, row=4).value = merged_disaggregation[2]
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + idx, row=4).alignment = Alignment(
                horizontal='center')
            merged_disaggregations_map[merged_disaggregation[1]] = DISAGGREGATION_COLUMN_START + idx

        # Merge disaggregation values
        merged_disaggregation_values = list()
        merged_disaggregation_values_map = dict()
        for sheet in self.sheets:
            column = DISAGGREGATION_COLUMN_START
            while (True):
                if sheet.cell(column=column, row=4).value and sheet.cell(column=column, row=4).value.find("#indicator+type+") > -1:
                    column += 1
                    continue
                if sheet.cell(column=column, row=4).value == None or sheet.cell(column=column, row=4).value.find("#indicator+value+") < 0:
                    break
                if (
                        sheet.cell(column=column, row=1).value,
                        sheet.cell(column=column, row=2).value,
                        sheet.cell(column=column, row=3).value,
                        sheet.cell(column=column, row=4).value
                ) not in merged_disaggregation_values:
                    merged_disaggregation_values.append(
                        (
                            sheet.cell(column=column, row=1).value,
                            sheet.cell(column=column, row=2).value,
                            sheet.cell(column=column, row=3).value,
                            sheet.cell(column=column, row=4).value
                        )
                    )
                column += 1

        # Prepare new column headers for disaggregation values
        for idx, merged_disaggregation_value in enumerate(merged_disaggregation_values):
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + idx, row=1).value = merged_disaggregation_value[0]
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + idx, row=1).font = Font(bold=True)
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + idx, row=1).alignment = Alignment(horizontal='center')
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + idx, row=2).value = merged_disaggregation_value[1]
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + idx, row=2).alignment = Alignment(horizontal='center')
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + idx, row=3).value = merged_disaggregation_value[2]
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + idx, row=3).alignment = Alignment(horizontal='center')
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + idx, row=4).value = merged_disaggregation_value[3]
            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + idx, row=4).alignment = Alignment(horizontal='center')
            merged_disaggregation_values_map[merged_disaggregation_value[1]] = DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + idx

        # Add Total column at the very end
        merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + len(merged_disaggregation_values), row=1).value = "Total"
        merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + len(merged_disaggregation_values), row=1).font = Font(bold=True)
        merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + len(merged_disaggregation_values), row=1).alignment = Alignment(horizontal='center')

        # Merge data from all sheets
        merged_row = INDICATOR_DATA_ROW_START
        for sheet in self.sheets:
            # Count columns
            max_columns = 1
            while(True):
                # Stop condition
                if sheet.cell(column=max_columns, row=1).value == None:
                    max_columns -= 1
                    break
                max_columns += 1
            sheet_row = INDICATOR_DATA_ROW_START
            while(True):
                # Stop row condition
                if sheet.cell(column=1, row=sheet_row).value == None:
                    break
                for column in range(max_columns):
                    column += 1
                    # If value less then DISAGGREGATION_COLUMN_START just copy data
                    if column < DISAGGREGATION_COLUMN_START:
                        merged_sheet.cell(column=column, row=merged_row).value = sheet.cell(column=column, row=sheet_row).value
                    else:
                        # Total has None value as a type
                        if sheet.cell(column=column, row=4).value == None:
                            merged_sheet.cell(column=DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + len(merged_disaggregation_values),
                                              row=merged_row).value = sheet.cell(column=column, row=sheet_row).value
                        # If column is disaggregation type
                        elif sheet.cell(column=column, row=4).value.find("#indicator+type+") > -1:
                            merged_sheet.cell(column=merged_disaggregations_map[sheet.cell(column=column, row=2).value], row=merged_row).value = sheet.cell(column=column, row=sheet_row).value
                        # If column is disaggregation value
                        elif sheet.cell(column=column, row=4).value.find("#indicator+value+") > -1:
                            merged_sheet.cell(column=merged_disaggregation_values_map[sheet.cell(column=column, row=2).value], row=merged_row).value = sheet.cell(column=column, row=sheet_row).value

                merged_row += 1
                sheet_row += 1
            # Remove basic sheets
            self.wb.remove_sheet(sheet)

        # Adjust columns width
        for col in merged_sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            merged_sheet.column_dimensions[column].width = adjusted_width

        # Add filters
        merged_sheet.auto_filter.ref = "A1:%s%d" % (
            get_column_letter(DISAGGREGATION_COLUMN_START + len(merged_disaggregations) + len(merged_disaggregation_values)),
            merged_row
        )

        # Lock first rows
        merged_sheet.freeze_panes = 'A%d' % INDICATOR_DATA_ROW_START

    def export_data(self):

        # Disaggregation types
        disaggregation_types_base = Disaggregation.objects.filter(response_plan__id=self.response_plan_id)

        # Split data into spreadsheets, combination of disaggregation_types_length with r-length tuples and no repeated elements
        disaggregation_types_list = list()
        for i in range(MAXIMUM_DISSAGREGATIONS_PER_INDICATOR + 1):
            disaggregation_types_list += list(itertools.combinations(disaggregation_types_base, i))

        # Duplicate spreadsheets base
        self.sheets = [self.sheet, ]
        for i in range(1, len(disaggregation_types_list)):
            self.sheets.append(self.duplicate_sheet(self.sheet))

        # Generate data per spreadsheet
        sheet_no = 0
        to_remove = list()
        for disaggregation_types in disaggregation_types_list:
            self.sheet = self.sheets[sheet_no]
            if not self.generate_sheet(disaggregation_types):
                to_remove.append(self.sheets[sheet_no])
            sheet_no += 1

        # Remove empty spreadsheets
        for s in to_remove:
            self.sheets.remove(s)
            self.wb.remove_sheet(s)

        if self.analysis:
            self.merge_sheets()

        filepath = SAVE_PATH + 'export.xlsx'
        self.wb.save(filepath)
        return filepath
