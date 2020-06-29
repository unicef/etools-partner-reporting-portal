import hashlib
import logging
import os
import tempfile

from django.http import HttpResponse
from django.utils import timezone

from easy_pdf.exceptions import PDFRenderingError
from easy_pdf.rendering import make_response, render_to_pdf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter
from unicef.exports.utilities import HTMLTableCell, HTMLTableHeader, PARTNER_PORTAL_DATE_FORMAT_EXCEL
from unicef.templatetags.pdf_extras import format_currency

logger = logging.getLogger(__name__)


def calc_cash_transfer_percentage(pd):
    return pd.funds_received_to_date_percentage / 100 if pd.budget else 0


class ProgrammeDocumentsXLSXExporter:

    def __init__(self, programme_documents):
        self.programme_documents = programme_documents
        filename = hashlib.sha256(';'.join([str(p.pk) for p in programme_documents]).encode('utf-8')).hexdigest()
        self.file_path = os.path.join(tempfile.gettempdir(), filename + '.xlsx')
        self.workbook = Workbook()
        self.worksheet = self.workbook.get_active_sheet()
        self.display_name = '[{:%a %-d %b %-H-%M-%S %Y}] {} Programme Document(s) Summary.xlsx'.format(
            timezone.now(), programme_documents.count()
        )

        self.header_style = NamedStyle(name="header")
        self.header_style.font = Font(bold=True, color='FFFFFF')
        self.header_style.fill = PatternFill("solid", fgColor="3195EE")
        self.header_style.alignment = Alignment(horizontal='center', vertical='justify')

    def fill_worksheet(self):
        self.worksheet.title = 'Programme Document(s) Summary'
        headers = [
            'Title',
            'PD/SSFA status',
            'Agreement',
            'Document Type',
            'Reference Number',
            'UNICEF Office(s)',
            'UNICEF Focal Point(s)',
            'Partner Focal Point(s)',
            'In response to an HRP',

            'Start date',
            'End date',
            'CSO contribution',
            'Total UNICEF cash',
            'Total UNICEF supplies',
            'Total Budget',
            'Cash Transfers to Date',
            'Cash Transfers to Date (%)',
        ]

        current_row = 1
        column_widths = []

        for column, header_text in enumerate(headers):
            column_widths.append(len(header_text))
            column += 1  # columns are not 0-indexed...
            cell = self.worksheet.cell(row=current_row, column=column, value=header_text)
            cell.style = self.header_style
        current_row += 1

        for pd in self.programme_documents:
            cash_transfer_percentage = calc_cash_transfer_percentage(pd)

            data_row = [
                (pd.title, None),
                (pd.get_status_display(), None),
                (pd.agreement, None),
                (pd.get_document_type_display(), None),
                (pd.reference_number, None),
                (pd.unicef_office, None),
                (', '.join([person.name for person in pd.unicef_focal_point.filter(active=True)]), None),
                (', '.join([person.name for person in pd.partner_focal_point.filter(active=True)]), None),
                (None, None),  # This field is not calculated anywhere yet
                (pd.start_date, PARTNER_PORTAL_DATE_FORMAT_EXCEL),
                (pd.end_date, PARTNER_PORTAL_DATE_FORMAT_EXCEL),
                (pd.cso_contribution, '#,##0.00_-[${} ]'.format(pd.cso_contribution_currency)),
                (pd.total_unicef_cash, '#,##0.00_-[${} ]'.format(pd.total_unicef_cash_currency)),
                (pd.in_kind_amount, '#,##0.00_-[${} ]'.format(pd.in_kind_amount_currency)),
                (pd.budget, '#,##0.00_-[${} ]'.format(pd.budget_currency)),
                (pd.funds_received_to_date, '#,##0.00_-[${} ]'.format(pd.funds_received_to_date_currency)),
                (cash_transfer_percentage, FORMAT_PERCENTAGE),
            ]

            if not len(headers) == len(data_row):
                raise Exception('Header and data row length mismatch!')

            for column, (cell_data, cell_format) in enumerate(data_row):
                try:
                    column_widths[column] = max(column_widths[column], len(cell_data) + 2)
                except TypeError:
                    column_widths[column] = max(column_widths[column], len(str(cell_data)) + 2)
                column += 1  # columns are not 0-indexed...
                cell = self.worksheet.cell(row=current_row, column=column, value=cell_data)
                if cell_format:
                    cell.number_format = cell_format

            current_row += 1

        for column, width in enumerate(column_widths):
            column += 1
            self.worksheet.column_dimensions[get_column_letter(column)].width = width

        self.workbook.save(self.file_path)

    def get_as_response(self):
        self.fill_worksheet()
        response = HttpResponse()
        response.content_type = self.worksheet.mime_type
        with open(self.file_path, 'rb') as content:
            response.write(content.read())
        self.cleanup()
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(self.display_name)
        return response

    def cleanup(self):
        os.remove(self.file_path)


class ProgrammeDocumentsPDFExporter:

    template_name = 'programme_documents_pdf_export.html'

    def __init__(self, programme_documents):
        self.programme_documents = programme_documents
        self.file_name = '[{:%a %-d %b %-H-%M-%S %Y}] {} Programme Document(s) Summary.pdf'.format(
            timezone.now(), programme_documents.count()
        )

    def get_context(self):
        context = {
            'title': 'Programme Document(s) Summary',
        }

        pd = self.programme_documents.first()
        partner_title = pd and pd.partner.title

        rows = [
            [
                HTMLTableHeader(partner_title, colspan=7),
            ],
            [
                HTMLTableHeader('Title'),
                HTMLTableHeader('General Info', colspan=4),
                HTMLTableHeader('Financial Info', colspan=2),
            ]
        ]

        for pd in self.programme_documents.order_by('id'):
            cash_transfer_percentage = calc_cash_transfer_percentage(pd)

            rows.append([
                HTMLTableCell(pd.title, rowspan=5),
                HTMLTableHeader('Agreement'),
                HTMLTableCell(pd.agreement),
                HTMLTableHeader('UNICEF Office(s)'),
                HTMLTableCell(pd.unicef_office),
                HTMLTableHeader('CSO contribution'),
                HTMLTableCell(format_currency(pd.cso_contribution, pd.cso_contribution_currency)),
            ])

            rows.append([
                HTMLTableHeader('Document Type'),
                HTMLTableCell(pd.get_document_type_display()),
                HTMLTableHeader('UNICEF Focal Point(s)'),
                HTMLTableCell(', '.join([person.name for person in pd.unicef_focal_point.filter(active=True)])),
                HTMLTableHeader('Total UNICEF cash'),
                HTMLTableCell(format_currency(pd.total_unicef_cash, pd.total_unicef_cash_currency))
            ])
            rows.append([
                HTMLTableHeader('Reference Number'),
                HTMLTableCell(pd.reference_number),
                HTMLTableHeader('Partner Focal Point(s)'),
                HTMLTableCell(', '.join([person.name for person in pd.partner_focal_point.filter(active=True)])),
                HTMLTableHeader('Total UNICEF supplies'),
                HTMLTableCell(format_currency(pd.in_kind_amount, pd.in_kind_amount_currency))
            ])
            rows.append([
                HTMLTableHeader('PD/SSFA status'),
                HTMLTableCell(pd.get_status_display()),
                HTMLTableHeader('Start Date'),
                HTMLTableCell(pd.start_date),
                HTMLTableHeader('Total Budget'),
                HTMLTableCell(format_currency(pd.budget, pd.budget_currency))
            ])
            rows.append([
                HTMLTableHeader('In response to an HRP'),
                HTMLTableCell(''),
                HTMLTableHeader('End Date'),
                HTMLTableCell(pd.end_date),
                HTMLTableHeader('Cash Transfers to Date'),
                HTMLTableCell('{} ({:.2f}%)'.format(
                    format_currency(pd.funds_received_to_date, pd.funds_received_to_date_currency),
                    cash_transfer_percentage,
                ))
            ])

        context['rows'] = rows
        context['programme_documents'] = self.programme_documents

        return context

    def get_as_response(self):
        try:
            pdf = render_to_pdf(self.template_name, self.get_context())
            response = make_response(pdf)
            response['Content-disposition'] = 'inline; filename="{}"'.format(self.file_name)
            return response
        except PDFRenderingError:
            logger.exception('Error trying to render PDF')
            return HttpResponse('Error trying to render PDF')
