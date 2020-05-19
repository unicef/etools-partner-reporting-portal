import logging

from django.utils import timezone

from babel.numbers import format_percent
from indicator.constants import ValueType
from indicator.models import Disaggregation, IndicatorBlueprint
from indicator.utilities import format_total_value_to_string
from openpyxl.utils import get_column_letter
from unicef.exports.annex_c_excel import ProgressReportsXLSXExporter
from unicef.exports.progress_reports import ProgressReportDetailPDFExporter
from unicef.exports.utilities import HTMLTableCell, HTMLTableHeader

logger = logging.getLogger(__name__)


class ReportableListXLSXExporter(ProgressReportsXLSXExporter):

    include_disaggregations = True
    export_to_single_sheet = True

    def __init__(self, reportables, **kwargs):
        self.reportables = reportables
        super(ReportableListXLSXExporter, self).__init__([], **kwargs)
        self.display_name = '[{:%a %-d %b %-H-%M-%S %Y}] Indicators Export.xlsx'.format(
            timezone.now()
        )

    def get_disaggregations(self, *args):
        return Disaggregation.objects.filter(
            reportable__in=self.reportables
        ).distinct()

    def write_reportables_to_current_sheet(self, reportables):
        current_row = self.write_header_to_current_sheet()

        disaggregation_column_map = self.write_disaggregation_headers_get_column_map()

        for reportable in reportables:
            reports = reportable.indicator_reports.order_by('-time_period_start')[:2]

            current_row = self.write_indicator_reports_to_current_sheet(
                current_row, reports, disaggregation_column_map
            )

        for column, width in enumerate(self.column_widths):
            column += 1
            self.current_sheet.column_dimensions[get_column_letter(column)].width = width

    def fill_workbook(self):
        reportables = self.reportables.prefetch_related(
            'indicator_reports'
        )
        if self.export_to_single_sheet:
            self.current_sheet.title = 'Indicators Export'
            self.write_reportables_to_current_sheet(reportables)
        else:
            for reportable in reportables:
                if not self.current_sheet.max_row == 1:
                    self.current_sheet = self.workbook.create_sheet()
                self.current_sheet.title = str(reportable)
                self.write_reportables_to_current_sheet([reportable])

        self.workbook.save(self.file_path)


class ReportableListPDFExporter(ProgressReportDetailPDFExporter):

    template_name = 'reportable_list_pdf_export.html'

    def __init__(self, reportables):
        self.reportables = reportables
        self.display_name = '[{:%a %-d %b %-H-%M-%S %Y}] List of Indicators'.format(
            timezone.now()
        )
        self.file_name = self.display_name + '.pdf'

    def get_reportable_header_table(self, reportable):
        is_percentage = reportable.blueprint.unit == IndicatorBlueprint.PERCENTAGE

        calculated_target = format_total_value_to_string(
            reportable.target,
            is_percentage=is_percentage,
            percentage_display_type="ratio" if reportable.blueprint.display_type == 'ratio' else None
        )

        if reportable.blueprint.display_type == 'percentage':
            calculated_baseline = "%.2f" % reportable.calculated_baseline * 100 if reportable.calculated_baseline * 100 != 0.0 else 0.0
        else:
            v = int(reportable.baseline.get(ValueType.VALUE, 0)) if reportable.baseline.get(ValueType.VALUE, 0) else 0
            d = int(reportable.baseline.get(ValueType.DENOMINATOR, 1)) if reportable.baseline.get(ValueType.DENOMINATOR, 1) else 1

            calculated_baseline = format_total_value_to_string(
                {
                    ValueType.VALUE: v,
                    ValueType.DENOMINATOR: d
                },
                is_percentage=is_percentage,
                percentage_display_type="ratio" if reportable.blueprint.display_type == 'ratio' else None
            )

        return [
            [
                HTMLTableHeader(reportable.blueprint.title, colspan=3, klass='section'),
            ],
            [
                HTMLTableHeader('Calculation method'),
                HTMLTableCell(reportable.blueprint.display_type, colspan=2),
            ],
            [
                HTMLTableHeader('Baseline'),
                HTMLTableCell(calculated_baseline, colspan=2),
            ],
            [
                HTMLTableHeader('Target'),
                HTMLTableCell(calculated_target, colspan=2),
            ],
            [
                HTMLTableHeader('Current Progress'),
                HTMLTableCell(format_percent(reportable.progress_percentage / 100, '#%'), colspan=2),
            ],
        ]

    def get_current_previous_location_data_table(self, current_data):
        previous_data = current_data.previous_location_data
        current_location_progress = format_total_value_to_string(
            current_data.disaggregation.get('()'),
            is_percentage=current_data.indicator_report.reportable.blueprint.unit == IndicatorBlueprint.PERCENTAGE,
            percentage_display_type="ratio" if current_data.indicator_report.reportable.blueprint.display_type == 'ratio' else None
        )
        previous_location_progress = format_total_value_to_string(
            previous_data.disaggregation.get('()'),
            is_percentage=previous_data.indicator_report.reportable.blueprint.unit == IndicatorBlueprint.PERCENTAGE,
            percentage_display_type="ratio" if current_data.indicator_report.reportable.blueprint.display_type == 'ratio' else None
        ) if previous_data else ''

        previous_time_period = previous_data.indicator_report.display_time_period if previous_data else ''
        previous_submission_date = previous_data.indicator_report.submission_date if previous_data else ''

        return [
            [
                HTMLTableHeader(current_data.location.title, colspan=3, klass='subsection'),
            ],
            [
                HTMLTableHeader('Reporting period'),
                HTMLTableHeader('Current {}'.format(current_data.indicator_report.display_time_period)),
                HTMLTableHeader('Previous {}'.format(previous_time_period)),
            ],
            [
                HTMLTableHeader('Total Progress'),
                HTMLTableCell(current_location_progress),
                HTMLTableCell(previous_location_progress),
            ],
            [
                HTMLTableHeader('Report submitted'),
                HTMLTableCell(current_data.indicator_report.submission_date),
                HTMLTableCell(previous_submission_date),
            ],
        ]

    def get_context(self):
        section_list = []

        for reportable in self.reportables:
            tables = [
                self.get_reportable_header_table(reportable)
            ]

            current_indicator_report = reportable.indicator_reports.order_by('-time_period_start').first()

            if not current_indicator_report:
                continue

            for location_data in current_indicator_report.indicator_location_data.all():
                tables.append(self.get_current_previous_location_data_table(
                    location_data
                ))

            section_data = {
                'tables': tables
            }

            section_list.append(section_data)

        context = {
            'title': self.display_name,
            'sections': section_list
        }

        return context
