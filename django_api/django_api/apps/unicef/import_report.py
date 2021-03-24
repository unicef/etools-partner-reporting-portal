import traceback
from datetime import datetime

from django.db import transaction

from core import common
from indicator.disaggregators import QuantityIndicatorDisaggregator, RatioIndicatorDisaggregator
from indicator.models import DisaggregationValue, IndicatorBlueprint, IndicatorLocationData
from openpyxl.reader.excel import load_workbook

from .models import ProgressReport

COLUMN_HASH_ID = 4
MAX_COLUMNS = 1000


class ProgressReportXLSXReader:

    def __init__(self, path, partner):
        self.wb = load_workbook(path)
        self.partner = partner

    def import_data(self):
        partner_contribution = None
        challenges = None
        proposed_way_forward = None
        pd_output_narratives = dict()

        for idx, self.sheet in enumerate(self.wb.worksheets):
            if self.sheet.title.lower() == 'readme':
                continue

            # Find "Location ID" column
            location_column_id = None
            for column in range(1, MAX_COLUMNS):
                if self.sheet.cell(row=COLUMN_HASH_ID, column=column).value == "#loc+id":
                    location_column_id = column
                    break
            if not location_column_id:
                return "Cannot find Location ID column"

            # Find "Total" column
            total_column_id = None
            for column in range(1, MAX_COLUMNS):
                if self.sheet.cell(row=1, column=column).value == "Total":
                    total_column_id = column
                    break
            if not total_column_id:
                return "Cannot find Total column"

            # Find first Disaggregation Value column
            dis_data_column_start_id = None
            for column in range(1, MAX_COLUMNS):
                if not self.sheet.cell(row=COLUMN_HASH_ID, column=column).value:
                    break
                if "#indicator+value" in self.sheet.cell(row=COLUMN_HASH_ID, column=column).value:
                    dis_data_column_start_id = column
                    break

            # Find "Progress" column
            progress_column_id = None
            for column in range(1, MAX_COLUMNS):
                if self.sheet.cell(row=COLUMN_HASH_ID, column=column).value == "#pr+id":
                    progress_column_id = column
                    break
            if not progress_column_id:
                return "Cannot find Progress ID column"

            # Other Info columns data retrieve
            # Partner contribution to date
            if not partner_contribution:
                partner_contribution = self.sheet.cell(
                    row=COLUMN_HASH_ID + 1, column=9).value

            # Challenges/bottlenecks in the reporting period
            if not challenges:
                challenges = self.sheet.cell(
                    row=COLUMN_HASH_ID + 1, column=11).value

            # Proposed way forward
            if not proposed_way_forward:
                proposed_way_forward = self.sheet.cell(
                    row=COLUMN_HASH_ID + 1, column=12).value

            # ... and assign if is QPR
            try:
                progress_id = self.sheet.cell(
                    row=COLUMN_HASH_ID + 1, column=progress_column_id).value
                pr = ProgressReport.objects.get(pk=progress_id)
            except ProgressReport.DoesNotExist:
                return "Cannot find Progress Report"

            # Iterate over rows and save disaggregation values
            for row in range(COLUMN_HASH_ID + 1, self.sheet.max_row):
                # If row is empty, end of sheet
                if not self.sheet.cell(row=row, column=1).value:
                    # Update Other Info sheet
                    if pr.report_type == common.QPR_TYPE:
                        pr.partner_contribution_to_date = partner_contribution
                        pr.challenges_in_the_reporting_period = challenges
                        pr.proposed_way_forward = proposed_way_forward
                        pr.save()
                    break

                # Get IndicatorLocationData ID
                try:
                    ild_id = str(
                        int(self.sheet.cell(row=row, column=location_column_id).value))

                    ind = IndicatorLocationData.objects.filter(pk=ild_id)

                    # Check if indicator has parent (UNICEF)
                    # If does, use parent to check partner
                    if self.partner and ind.filter(indicator_report__parent__isnull=False):
                        if not ind.filter(
                                indicator_report__parent__reportable__partner_activity_project_contexts__project__partner=self.partner
                        ).exists():
                            return "Parent of Indicator ID " \
                                + ild_id \
                                + " does not belong to partner " \
                                + str(self.partner)
                    # Check if Partner is allowed to modify data
                    elif self.partner and ind.filter(
                            indicator_report__progress_report__programme_document__partner=self.partner).count() == 0:
                        return "Indicator ID " + ild_id + " does not belong to partner " + str(self.partner)
                    indicator = IndicatorLocationData.objects.get(
                        pk=int(self.sheet.cell(
                            row=row, column=location_column_id).value)
                    )

                    # Check if Indicator Report is not able to submit anymore
                    if not indicator.indicator_report.can_import:
                        transaction.rollback()
                        return "Indicator in row {} is already submitted. Please remove row and try again.".format(
                            row,)

                except IndicatorLocationData.DoesNotExist:
                    return "Cannot find Indicator Location Data data for ID " \
                        + str(self.sheet.cell(row=row,
                                              column=location_column_id).value)

                blueprint = indicator.indicator_report.reportable.blueprint
                data = indicator.disaggregation

                if pr.report_type == common.QPR_TYPE:
                    narrative_assessment = self.sheet.cell(
                        row=row, column=19
                    ).value
                    llo = indicator.indicator_report.reportable.content_object

                    if llo.id not in pd_output_narratives \
                            and (narrative_assessment is not None and narrative_assessment != ''):
                        pd_output_narratives[llo.id] = narrative_assessment

                        indicator.indicator_report.narrative_assessment = narrative_assessment
                        indicator.indicator_report.save()

                        pr.indicator_reports.filter(reportable__lower_level_outputs=llo).update(narrative_assessment=narrative_assessment)

                # Prepare
                already_updated_row_value = False
                for column in range(dis_data_column_start_id if dis_data_column_start_id else total_column_id,
                                    total_column_id + 1):
                    try:
                        value = self.sheet.cell(row=row, column=column).value
                        # Check if value is present in cell
                        if value is not None:
                            # Evaluate ID of Disaggregation Type
                            dis_type_id = "()"
                            dis_type_value = self.sheet.cell(
                                row=2, column=column).value
                            if dis_type_value:
                                dis_type_value = sorted(
                                    list(map(int, str(dis_type_value).split(","))), key=int)
                                dis_type_id = str(tuple(dis_type_value))

                            if dis_type_id not in data:
                                # Check if data is proper disaggregation value
                                for dt in dis_type_value:
                                    if not DisaggregationValue.objects.filter(pk=dt).exists():
                                        transaction.rollback()
                                        return "Disaggregation {} does not exists".format(
                                            self.sheet.cell(row=4, column=column).value)
                                    # Check if filled disaggregation values
                                    # belongs to their type
                                    dv = DisaggregationValue.objects.get(pk=dt)
                                    if dv.disaggregation.id not in indicator.disaggregation_reported_on:
                                        transaction.rollback()
                                        return "Disaggregation {} does not belong to this Indicator".format(
                                            self.sheet.cell(row=4, column=column).value)
                                # Create value
                                data[dis_type_id] = dict()

                            already_updated_row_value = True

                            # Update values
                            if blueprint.unit == IndicatorBlueprint.NUMBER:
                                data[dis_type_id]["v"] = int(value)
                            else:
                                if isinstance(value, datetime):
                                    transaction.rollback()
                                    return "Value in column {}, row {} is Date Time. Please format row to Plain Text."\
                                        .format(self.sheet.cell(row=4, column=column).value, row)
                                values = value.split("/")
                                data[dis_type_id]["v"] = int(values[0])
                                data[dis_type_id]["d"] = int(values[1])
                        else:
                            # if value is not present, check if it should be
                            # all rows need to updated

                            # Evaluate ID of Disaggregation Type
                            dis_type_value = self.sheet.cell(
                                row=2, column=column).value
                            if dis_type_value:
                                dis_type_value = sorted(
                                    list(map(int, str(dis_type_value).split(","))), key=int)
                            if dis_type_value:
                                if len(dis_type_value) == indicator.level_reported:
                                    # Check if data is proper disaggregation
                                    # value
                                    for dt in dis_type_value:
                                        dv = DisaggregationValue.objects.get(
                                            pk=dt)
                                        if dv.disaggregation.id in indicator.disaggregation_reported_on and \
                                                already_updated_row_value:
                                            transaction.rollback()
                                            return "Please fulfill required value to column {}, row {}"\
                                                .format(self.sheet.cell(row=4, column=column).value, row)

                    except Exception:
                        traceback.print_exc()
                        transaction.rollback()
                        return "Cannot assign disaggregation value to column {}, row {}"\
                            .format(self.sheet.cell(row=4, column=column).value, row)

                indicator.disaggregation = data
                indicator.save()

                if blueprint.unit == IndicatorBlueprint.NUMBER:
                    QuantityIndicatorDisaggregator.post_process(indicator)

                if blueprint.unit == IndicatorBlueprint.PERCENTAGE:
                    RatioIndicatorDisaggregator.post_process(indicator)

        return
