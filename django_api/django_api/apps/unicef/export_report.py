import itertools

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.db.models import Count

from core import common

from indicator.models import DisaggregationValue, IndicatorBlueprint
from indicator.constants import ValueType
from indicator.utilities import convert_string_number_to_float

PATH = settings.BASE_DIR + "/apps/unicef/templates/excel/hr_export.xlsx"
SAVE_PATH = '/tmp/'

DISAGGREGATION_COLUMN_START = 43
INDICATOR_DATA_ROW_START = 5
MAXIMUM_DISAGGREGATIONS_PER_INDICATOR = 3
REQUIRED_FILL = PatternFill(fill_type='solid', start_color='FFE5A479', end_color='FFE5A479')
NO_FILL = PatternFill(fill_type=None)


class ProgressReportXLSXExporter:

    analysis = False

    def __init__(self, progress_report, analysis=None):
        self.wb = load_workbook(PATH)
        self.sheet = self.wb.get_sheet_by_name('PR Template')
        self.progress_report = progress_report
        if analysis is not None:
            self.analysis = analysis
        self.sheets = [self.wb.get_sheet_by_name('README'), self.sheet, ]
        self.disaggregations_start_column = DISAGGREGATION_COLUMN_START

        self.bold_center_style = NamedStyle(name="Bold and Center")
        self.bold_center_style.font = Font(bold=True)
        self.bold_center_style.alignment = Alignment(horizontal='center')

    def duplicate_sheet(self, sheet):
        return self.wb.copy_worksheet(sheet)

    def fill_sheet(self, disaggregation_types, indicators):
        # Setup a title (limited to 30 chars via excel requirements)
        self.sheet.title = ", ".join(
            [dt.name for dt in disaggregation_types]
        )[:30] if disaggregation_types else "None"

        # Generate disaggregation types columns
        # it holds column number for given Disaggregation Type ID
        disaggregation_types_map = dict()
        for idx, dt in enumerate(disaggregation_types):
            current_column = self.disaggregations_start_column + idx

            name_cell = self.sheet.cell(
                row=1, column=current_column, value=dt.name)
            name_cell.alignment = Alignment(horizontal='center')
            name_cell.font = Font(bold=True)

            id_cell = self.sheet.cell(
                row=2, column=current_column, value=dt.id)
            id_cell.alignment = Alignment(horizontal='center')

            type_cell = self.sheet.cell(
                row=4, column=current_column, value="#indicator+type+" + dt.name
            )
            type_cell.alignment = Alignment(horizontal='center')

            disaggregation_types_map[dt.id] = current_column

        # Hide non QPR columns
        if self.progress_report.report_type != common.QPR_TYPE:
            for col in ['L', 'M', 'O', 'P', 'Q']:
                self.sheet.column_dimensions[col].hidden = True

        # Prepare Disaggregation Values for given disaggregation types
        disaggregation_values = list()
        disaggregation_values_list = list()
        disaggregation_values_map = dict()
        disaggregation_values_by_type = dict()
        for disaggregation_type in disaggregation_types:
            disaggregation_values_base = DisaggregationValue.objects.filter(
                disaggregation=disaggregation_type
            ).order_by('value')
            disaggregation_values_list.append(disaggregation_values_base)
            for dv in disaggregation_values_base:
                disaggregation_values_by_type[dv.id] = disaggregation_type.name

        # Create all possible combinations (max. 3 items per combination) for
        # disaggregation values
        for combination_items in range(1, MAXIMUM_DISAGGREGATIONS_PER_INDICATOR + 1):
            for combinations in itertools.combinations(
                    disaggregation_values_list, combination_items):
                disaggregation_values += list(itertools.product(*combinations))

        # Generate disaggregation values columns
        for idx, dvs in enumerate(disaggregation_values):
            current_column = self.disaggregations_start_column + \
                len(disaggregation_types) + idx

            cell = self.sheet.cell(
                row=1, column=current_column, value=" + ".join([dv.value for dv in dvs]))
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)

            self.sheet.cell(row=2, column=current_column,
                            value=", ".join([str(dv.id) for dv in dvs]))
            self.sheet.cell(
                row=3, column=current_column, value=" + ".join([disaggregation_values_by_type[dv.id] for dv in dvs])
            )
            self.sheet.cell(
                row=4, column=current_column, value="#indicator+value+" + str("+".join([dv.value for dv in dvs]))
            )

            disaggregation_values_map[
                ", ".join([str(dv.id) for dv in dvs])] = current_column

        totals_column = self.disaggregations_start_column + \
            len(disaggregation_types) + len(disaggregation_values)
        totals_header_cell = self.sheet.cell(
            row=1, column=totals_column, value="Total")
        totals_header_cell.alignment = Alignment(horizontal='center')

        disaggregation_values_map["()"] = totals_column

        # Each row is one location per indicator
        start_row_id = INDICATOR_DATA_ROW_START

        all_indicators = list()
        # Dual Reporting
        for indicator in indicators:
            all_indicators.append(indicator)
            for children_indicator in indicator.children.all():
                all_indicators.append(children_indicator)

        for indicator in all_indicators:
            locations_data = indicator.indicator_location_data
            for location_data in locations_data.all():

                if indicator.display_type == IndicatorBlueprint.RATIO:
                    indicator_target = "{} / {}".format(
                        indicator.reportable.target['v'], indicator.reportable.target['d']
                    )
                else:
                    try:
                        indicator_target = convert_string_number_to_float(
                            indicator.reportable.calculated_target)
                    except ValueError:
                        indicator_target = indicator.reportable.calculated_target
                if indicator.display_type == IndicatorBlueprint.PERCENTAGE:
                    indicator_target = indicator_target * 100

                if indicator.is_percentage:
                    achievement_in_reporting_period = indicator.total.get(
                        ValueType.CALCULATED, 0)
                    total_cumulative_progress = indicator.reportable.achieved.get(
                        ValueType.CALCULATED, 0
                    )
                else:
                    achievement_in_reporting_period = indicator.total.get(
                        ValueType.VALUE, 0)
                    total_cumulative_progress = indicator.reportable.achieved.get(
                        ValueType.VALUE, 0
                    )

                self.sheet.cell(row=start_row_id, column=1).value = \
                    self.progress_report.programme_document.partner.title
                self.sheet.cell(row=start_row_id, column=2).value = \
                    location_data.location.gateway.country.name
                self.sheet.cell(row=start_row_id, column=3).value = \
                    self.progress_report.programme_document.reference_number
                self.sheet.cell(row=start_row_id, column=4).value = \
                    self.progress_report.programme_document.title
                self.sheet.cell(row=start_row_id, column=5).value = \
                    self.progress_report.get_reporting_period()
                self.sheet.cell(row=start_row_id, column=6).value = \
                    self.progress_report.get_status_display()
                self.sheet.cell(row=start_row_id, column=7).value = \
                    self.progress_report.due_date
                self.sheet.cell(row=start_row_id, column=8).value = \
                    self.progress_report.submission_date
                self.sheet.cell(row=start_row_id, column=9).value = \
                    self.progress_report.partner_contribution_to_date
                self.sheet.cell(row=start_row_id, column=9).fill = \
                    REQUIRED_FILL
                self.sheet.cell(row=start_row_id, column=10).value = \
                    self.progress_report.programme_document.funds_received_to_date
                self.sheet.cell(row=start_row_id, column=11).value = \
                    self.progress_report.challenges_in_the_reporting_period
                self.sheet.cell(row=start_row_id, column=11).fill = \
                    REQUIRED_FILL
                self.sheet.cell(row=start_row_id, column=12).value = \
                    self.progress_report.proposed_way_forward
                self.sheet.cell(row=start_row_id, column=12).fill = \
                    REQUIRED_FILL
                self.sheet.cell(row=start_row_id, column=13).value = \
                    self.progress_report.submitted_by.display_name if \
                    self.progress_report.submitted_by else ''

                attachments = self.progress_report.attachments.all()
                other_attachments = attachments.filter(type="Other")

                self.sheet.cell(row=start_row_id, column=14).value = \
                    attachments.filter(type="FACE").first().file.url if attachments.filter(type="FACE").exists() else ''

                if other_attachments.exists():
                    if other_attachments.count() == 1:
                        self.sheet.cell(row=start_row_id, column=15).value = other_attachments.first().file.url
                        self.sheet.cell(row=start_row_id, column=16).value = ''
                    elif other_attachments.count() > 1:
                        self.sheet.cell(row=start_row_id, column=15).value = other_attachments.first().file.url
                        self.sheet.cell(row=start_row_id, column=16).value = other_attachments.last().file.url
                else:
                    self.sheet.cell(row=start_row_id, column=15).value = ''
                    self.sheet.cell(row=start_row_id, column=16).value = ''

                self.sheet.cell(row=start_row_id, column=17).value = \
                    indicator.reportable.content_object.title

                self.sheet.cell(row=start_row_id, column=18).value = \
                    indicator.get_overall_status_display()
                self.sheet.cell(row=start_row_id, column=19).value = \
                    indicator.narrative_assessment
                self.sheet.cell(row=start_row_id, column=19).fill = \
                    REQUIRED_FILL
                self.sheet.cell(row=start_row_id, column=20).value = \
                    indicator.reportable.blueprint.title
                self.sheet.cell(row=start_row_id, column=21).value = \
                    indicator.display_type
                self.sheet.cell(row=start_row_id, column=22).value = \
                    indicator_target
                self.sheet.cell(row=start_row_id, column=23).value = \
                    indicator.calculation_formula_across_locations
                self.sheet.cell(row=start_row_id, column=24).value = \
                    indicator.calculation_formula_across_periods
                self.sheet.cell(row=start_row_id, column=25).value = \
                    location_data.previous_location_progress_value

                # Iterate over location admin references:
                location = location_data.location
                while True:
                    admin_level = location.gateway.admin_level
                    # TODO: secure in case of wrong location data
                    admin_level = min(admin_level, 5)
                    self.sheet.cell(row=start_row_id, column=25 +
                                    admin_level * 2).value = location.title
                    self.sheet.cell(row=start_row_id, column=25 +
                                    admin_level * 2 - 1).value = location.gateway.name

                    if location.parent:
                        location = location.parent
                    else:
                        break

                self.sheet.cell(row=start_row_id, column=36).value = \
                    indicator.reportable.numerator_label
                self.sheet.cell(row=start_row_id, column=37).value = \
                    indicator.reportable.denominator_label
                self.sheet.cell(row=start_row_id, column=38).value = \
                    achievement_in_reporting_period
                self.sheet.cell(row=start_row_id, column=39).value = \
                    total_cumulative_progress
                self.sheet.cell(row=start_row_id, column=40).value = \
                    self.progress_report.id
                self.sheet.cell(row=start_row_id, column=41).value = \
                    indicator.id
                self.sheet.cell(row=start_row_id, column=42).value = \
                    location_data.id

                # Check location item disaggregation type
                for reported_disaggregation_type in location_data.disaggregation_reported_on:
                    if reported_disaggregation_type in disaggregation_types_map:
                        self.sheet.cell(
                            row=start_row_id,
                            column=disaggregation_types_map[reported_disaggregation_type]).value = "X"

                # Check location item values
                blueprint = location_data.indicator_report.reportable.blueprint

                for dk, dv in disaggregation_values_map.items():
                    self.sheet.cell(row=start_row_id, column=dv).fill = REQUIRED_FILL

                for k, v in location_data.disaggregation.items():
                    if k == "()":
                        self.sheet.cell(
                            row=start_row_id,
                            column=disaggregation_values_map['()']).value = v['v'] \
                            if blueprint.unit == IndicatorBlueprint.NUMBER else "{}/{}".format(v['v'], v['d'])

                        # De-highlight total column if the indicator is disaggregated
                        if disaggregation_values_list:
                            self.sheet.cell(
                                row=start_row_id,
                                column=disaggregation_values_map['()']).fill = NO_FILL

                    else:
                        for dk, dv in disaggregation_values_map.items():
                            if dk == "()":
                                continue
                            if sorted(list(eval(k))) == sorted(
                                    list(int(k) for k in dk.split(","))):
                                self.sheet.cell(
                                    row=start_row_id, column=dv).value = v['v'] \
                                    if blueprint.unit == IndicatorBlueprint.NUMBER \
                                    else "{}/{}".format(v['v'], v['d'])

                                # De-highlight any subtotal disaggregation data column
                                if len(dk.split(",")) != len(disaggregation_types):
                                    self.sheet.cell(row=start_row_id, column=dv).fill = NO_FILL

                start_row_id += 1

        # Lock first rows
        self.sheet.freeze_panes = 'A%d' % INDICATOR_DATA_ROW_START

        # Merge Other Info columns, since they are unique per Progress Report, not per Location Data
        # Partner contribution to date
        self.sheet.merge_cells(start_row=INDICATOR_DATA_ROW_START,
                               start_column=9, end_row=start_row_id, end_column=9)
        # Challenges/bottlenecks in the reporting period
        self.sheet.merge_cells(start_row=INDICATOR_DATA_ROW_START,
                               start_column=11, end_row=start_row_id, end_column=11)
        # Proposed way forward
        self.sheet.merge_cells(start_row=INDICATOR_DATA_ROW_START,
                               start_column=12, end_row=start_row_id, end_column=12)

        return True

    def merge_sheets(self):
        """
        1. Create new sheet for Analysis
        2. Copy generic columns (same for all)
        3. Merge disaggregation types columns
        4. Merge disaggregation values columns
        5. Add Total column at the very end of sheet
        6. Iterate over basic sheets and merge data into Analysis sheet
        7. Remove basic sheets
        """

        # Create new sheet
        merged_sheet = self.wb.create_sheet(title="Analysis")

        # Copy generic columns
        for i in range(1, self.disaggregations_start_column):
            merged_sheet.cell(
                column=i, row=1, value=self.sheet.cell(column=i, row=1).value
            ).style = self.bold_center_style

            merged_sheet.cell(
                column=i, row=2, value=self.sheet.cell(column=i, row=2).value
            ).alignment = Alignment(horizontal='center')

            merged_sheet.cell(
                column=i, row=3, value=self.sheet.cell(column=i, row=3).value
            ).alignment = Alignment(horizontal='center')

            merged_sheet.cell(
                column=i, row=4, value=self.sheet.cell(column=i, row=4).value
            ).alignment = Alignment(horizontal='center')

        # Merge disaggregation types
        merged_disaggregations = list()
        merged_disaggregations_map = dict()

        for sheet in self.sheets:
            column = self.disaggregations_start_column
            while True:
                if sheet.cell(column=column, row=4).value is None or sheet.cell(
                        column=column, row=4).value.find("#indicator+type+") < 0:
                    break
                if (
                    sheet.cell(column=column, row=1).value,
                    sheet.cell(column=column, row=2).value,
                    sheet.cell(column=column, row=4).value
                ) not in merged_disaggregations:
                    merged_disaggregations.append(
                        (
                            sheet.cell(column=column, row=1).value,
                            sheet.cell(column=column, row=2).value,
                            sheet.cell(column=column, row=4).value
                        )
                    )
                column += 1

        # Prepare new column headers for disaggregation types
        for idx, (
            disaggregation_name,
            disaggregation_id,
            disaggregation_label,
        ) in enumerate(merged_disaggregations):
            current_column = self.disaggregations_start_column + idx
            name_cell = merged_sheet.cell(
                column=current_column, row=1, value=disaggregation_name)
            name_cell.style = self.bold_center_style

            id_cell = merged_sheet.cell(
                column=current_column, row=2, value=disaggregation_id)
            id_cell.alignment = Alignment(horizontal='center')

            label_cell = merged_sheet.cell(
                column=current_column, row=4, value=disaggregation_label)
            label_cell.alignment = Alignment(horizontal='center')

            merged_disaggregations_map[disaggregation_id] = current_column

        # Merge disaggregation values
        merged_disaggregation_values = list()
        merged_disaggregation_values_map = dict()
        for sheet in self.sheets:
            column = self.disaggregations_start_column
            while True:
                if sheet.cell(column=column, row=4).value and sheet.cell(
                        column=column, row=4).value.find("#indicator+type+") > -1:
                    column += 1
                    continue
                if sheet.cell(column=column, row=4).value is None or sheet.cell(
                        column=column, row=4).value.find("#indicator+value+") < 0:
                    break
                if (
                        sheet.cell(column=column, row=1).value,
                        sheet.cell(column=column, row=2).value,
                        sheet.cell(column=column, row=3).value,
                        sheet.cell(column=column, row=4).value
                ) not in merged_disaggregation_values:
                    merged_disaggregation_values.append(
                        (
                            sheet.cell(column=column, row=1).value,
                            sheet.cell(column=column, row=2).value,
                            sheet.cell(column=column, row=3).value,
                            sheet.cell(column=column, row=4).value
                        )
                    )
                column += 1

        # Prepare new column headers for disaggregation values
        for idx, (
            disaggregation_value,
            disaggregation_id,
            disaggregation_name,
            disaggregation_label,
        ) in enumerate(merged_disaggregation_values):
            current_column = self.disaggregations_start_column + \
                len(merged_disaggregations) + idx
            value_cell = merged_sheet.cell(
                column=current_column, row=1, value=disaggregation_value)
            value_cell.style = self.bold_center_style

            id_cell = merged_sheet.cell(
                column=current_column, row=2, value=disaggregation_id)
            id_cell.alignment = Alignment(horizontal='center')

            name_cell = merged_sheet.cell(
                column=current_column, row=3, value=disaggregation_name)
            name_cell.alignment = Alignment(horizontal='center')

            label_cell = merged_sheet.cell(
                column=current_column, row=4, value=disaggregation_label)
            label_cell.alignment = Alignment(horizontal='center')

            merged_disaggregation_values_map[
                disaggregation_id] = current_column

        # Add Total column at the very end
        totals_column = self.disaggregations_start_column + \
            len(merged_disaggregations) + \
            len(merged_disaggregation_values)

        totals_header_cell = merged_sheet.cell(
            column=totals_column, row=1, value="Total")
        totals_header_cell.style = self.bold_center_style

        # Merge data from all sheets
        merged_row = INDICATOR_DATA_ROW_START
        for sheet in self.sheets:
            # Count columns
            max_columns = 1
            while True:
                # Stop condition
                if sheet.cell(column=max_columns, row=1).value is None:
                    max_columns -= 1
                    break
                max_columns += 1
            sheet_row = INDICATOR_DATA_ROW_START
            while True:
                # Stop row condition
                if sheet.cell(column=1, row=sheet_row).value is None:
                    break
                for column in range(max_columns):
                    column += 1
                    # If value less then self.disaggregations_start_column just copy
                    # data
                    if column < self.disaggregations_start_column:
                        merged_sheet.cell(
                            column=column, row=merged_row).value = sheet.cell(
                            column=column, row=sheet_row).value
                    else:
                        # Total has None value as a type
                        if sheet.cell(column=column, row=4).value is None:
                            col = self.disaggregations_start_column + \
                                len(merged_disaggregations) + \
                                len(merged_disaggregation_values)
                            merged_sheet.cell(
                                column=col, row=merged_row
                            ).value = sheet.cell(column=column, row=sheet_row).value
                        # If column is disaggregation type
                        elif sheet.cell(column=column, row=4).value.find("#indicator+type+") > -1:
                            merged_sheet.cell(
                                column=merged_disaggregations_map[
                                    sheet.cell(column=column, row=2).value],
                                row=merged_row
                            ).value = sheet.cell(column=column, row=sheet_row).value
                        # If column is disaggregation value
                        elif sheet.cell(column=column, row=4).value.find("#indicator+value+") > -1:
                            merged_sheet.cell(
                                column=merged_disaggregation_values_map[
                                    sheet.cell(column=column, row=2).value],
                                row=merged_row
                            ).value = sheet.cell(column=column, row=sheet_row).value

                merged_row += 1
                sheet_row += 1
            # Remove basic sheets
            self.wb.remove_sheet(sheet)

        # Adjust columns width
        for col in merged_sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except BaseException:
                    pass
            adjusted_width = max_length + 2
            merged_sheet.column_dimensions[column].width = adjusted_width

        # Add filters
        merged_sheet.auto_filter.ref = "A1:%s%d" % (
            get_column_letter(
                self.disaggregations_start_column +
                len(merged_disaggregations) +
                len(merged_disaggregation_values)),
            merged_row
        )

        # Lock first rows
        merged_sheet.freeze_panes = 'A%d' % INDICATOR_DATA_ROW_START

    def export_data(self):
        # Prepare list of unique disaggregations for choosed indicators
        self.indicators = self.progress_report.indicator_reports.filter(
            reportable__is_cluster_indicator=False)
        disaggregation_types_base_set = set()
        for i in self.indicators:
            for d in i.reportable.disaggregations.all():
                disaggregation_types_base_set.add(d)
        disaggregation_types_base = list(disaggregation_types_base_set)

        # Split data into spreadsheets, combination of
        # disaggregation_types_length with r-length tuples and no repeated
        # elements
        disaggregation_types_list = list()
        for i in range(MAXIMUM_DISAGGREGATIONS_PER_INDICATOR + 1):
            disaggregation_types_list += list(
                itertools.combinations(
                    disaggregation_types_base, i))

        # Generate data per spreadsheet

        sheet_no = 1
        to_remove = list()
        for disaggregation_types in disaggregation_types_list:
            # Check if indicators belongs to this disaggregation_types
            # Skip if not
            indicators = self.indicators
            # indicators = IndicatorReport.objects.filter(pk__in=[i.id for i in indicators])
            if disaggregation_types:
                indicators = indicators.annotate(
                    count=Count('reportable__disaggregations')).filter(
                    count=len(disaggregation_types))
                for dt in disaggregation_types:
                    indicators = indicators.filter(
                        reportable__disaggregations=dt)
            else:
                indicators = indicators.filter(
                    reportable__disaggregations__isnull=True)

            indicators = indicators.order_by('reportable__id')
            if not indicators:
                continue

            self.sheets.append(self.duplicate_sheet(self.sheets[1]))
            sheet_no += 1
            self.sheet = self.sheets[sheet_no]

            if not self.fill_sheet(disaggregation_types, indicators):
                to_remove.append(self.sheets[sheet_no])
                sheet_no -= 1

        to_remove.append(self.sheets[1])

        # Remove empty spreadsheets
        for s in to_remove:
            # Spreadsheet need at least 2 sheets
            if len(self.sheets) > 2:
                self.sheets.remove(s)
                self.wb.remove_sheet(s)

        if self.analysis:
            self.merge_sheets()

        report_name = self.progress_report.report_type + str(self.progress_report.report_number)
        ref_num = self.progress_report.programme_document.reference_number.split('/')[-1]

        file_path = SAVE_PATH + f'{report_name}_{ref_num}.xlsx'
        self.wb.save(file_path)
        return file_path
