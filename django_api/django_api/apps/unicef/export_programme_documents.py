import hashlib
import logging
import os
import tempfile

from babel.numbers import format_currency
from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone
from django.utils.translation import to_locale, get_language
from easy_pdf.exceptions import PDFRenderingError
from easy_pdf.rendering import render_to_pdf, make_response
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD, FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)


class ProgrammeDocumentsXLSXExporter:

    def __init__(self, programme_documents):
        self.programme_documents = programme_documents
        filename = hashlib.sha256(';'.join([str(p.pk) for p in programme_documents]).encode('utf-8')).hexdigest()
        self.file_path = os.path.join(tempfile.gettempdir(), filename + '.xlsx')
        self.workbook = Workbook()
        self.worksheet = self.workbook.get_active_sheet()
        self.display_name = '[{:%a %-d %b %-H-%M-%S %Y}] {} Programme Document(s) Summary.xlsx'.format(
            timezone.now(), programme_documents.count()
        )

        self.header_style = NamedStyle(name="header")
        self.header_style.font = Font(bold=True, color='FFFFFF')
        self.header_style.fill = PatternFill("solid", fgColor="3195EE")
        self.header_style.alignment = Alignment(horizontal='center', vertical='justify')

    def fill_worksheet(self):
        self.worksheet.title = 'Programme Document(s) Summary'
        headers = [
            'PD/SSFA ToR ref. #',
            'PD/SSFA status',
            'Start date',
            'End date',
            'CSO contribution',
            'UNICEF cash',
            'UNICEF supplies',
            'Planned Budget',
            'Cash Transfers to Date ($)',
            'Cash Transfers to Date (%)',
        ]

        current_row = 1

        for column, header_text in enumerate(headers):
            column += 1  # columns are not 0-indexed...
            cell = self.worksheet.cell(row=current_row, column=column, value=header_text)
            cell.style = self.header_style
            self.worksheet.column_dimensions[get_column_letter(column)].width = len(header_text) + 5
        current_row += 1

        for pd in self.programme_documents:
            if pd.budget:
                funds_received_to_date_percentage = pd.funds_received_to_date / pd.budget
            else:
                funds_received_to_date_percentage = 0

            data_row = [
                (pd.title, None),
                (pd.get_status_display(), None),
                (pd.start_date, None),
                (pd.end_date, None),
                (pd.cso_contribution, FORMAT_CURRENCY_USD),
                (pd.total_unicef_cash, FORMAT_CURRENCY_USD),
                (pd.in_kind_amount, FORMAT_CURRENCY_USD),
                (pd.budget, FORMAT_CURRENCY_USD),
                (pd.funds_received_to_date, FORMAT_CURRENCY_USD),
                (funds_received_to_date_percentage, FORMAT_PERCENTAGE),
            ]

            for column, (cell_data, cell_format) in enumerate(data_row):
                column += 1  # columns are not 0-indexed...
                cell = self.worksheet.cell(row=current_row, column=column, value=cell_data)
                if cell_format:
                    cell.number_format = cell_format
            current_row += 1

        self.workbook.save(self.file_path)

    def get_as_response(self):
        self.fill_worksheet()
        response = HttpResponse()
        response.content_type = self.worksheet.mime_type
        with open(self.file_path, 'rb') as content:
            response.write(content.read())
        self.cleanup()
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(self.display_name)
        return response

    def cleanup(self):
        os.remove(self.file_path)


class ProgrammeDocumentsPDFExporter:

    template_name = 'programme_documents_pdf_export.html'

    def __init__(self, programme_documents):
        self.programme_documents = programme_documents
        self.display_name = '[{:%a %-d %b %-H-%M-%S %Y}] {} Programme Document(s) Summary.pdf'.format(
            timezone.now(), programme_documents.count()
        )

    def get_context(self):
        context = {
            'title': 'Programme Document(s) Summary',
            'headers': [
                ('PD/SSFA ToR ref. #', 20),
                ('PD/SSFA status', 10),
                ('Start date', 10),
                ('End date', 10),
                ('CSO contribution', 10),
                ('UNICEF cash', 10),
                ('UNICEF supplies', 10),
                ('Planned Budget', 10),
                ('Cash Transfers to Date (%)', 10),
            ]
        }

        total_percentage_width = sum([h[1] for h in context['headers']])
        if not total_percentage_width == 100:
            raise Exception('Percentage widths must add up to 100, currently: {}'.format(total_percentage_width))

        data_rows = []

        locale = to_locale(get_language())

        for pd in self.programme_documents:
            if pd.budget:
                funds_received_to_date_percentage = pd.funds_received_to_date / pd.budget
            else:
                funds_received_to_date_percentage = 0

            data_rows.append([
                pd.title,
                pd.get_status_display,
                pd.start_date.strftime(settings.PRINT_DATA_FORMAT),
                pd.end_date.strftime(settings.PRINT_DATA_FORMAT),
                format_currency(pd.cso_contribution, pd.cso_contribution_currency, locale=locale),
                format_currency(pd.total_unicef_cash, pd.total_unicef_cash_currency, locale=locale),
                format_currency(pd.in_kind_amount, pd.in_kind_amount_currency, locale=locale),
                format_currency(pd.budget, pd.budget_currency, locale=locale),
                '{} ({}%)'.format(
                    format_currency(pd.funds_received_to_date, pd.funds_received_to_date_currency, locale=locale),
                    int(round(funds_received_to_date_percentage * 100, 0)),
                ),
            ])

        context['data_rows'] = data_rows

        return context

    def get_as_response(self):
        try:
            pdf = render_to_pdf(self.template_name, self.get_context())
            return make_response(pdf, self.display_name)
        except PDFRenderingError:
            logger.exception('Error trying to render PDF')
            return HttpResponse('Error trying to render PDF')
